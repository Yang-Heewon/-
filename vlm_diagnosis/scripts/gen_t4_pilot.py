"""Build the T4 pilot manifest v2 (ScreenQA content + auto-generated location questions).

This is the corrected (v2) generator. It supersedes the v1 session-scratchpad script
(select_t4.py) and fixes every issue from the 2026-08-15 formal audit:

  A. Nine v1 location questions failed visual audit (full-row/spinner bboxes, UI-state
     answers, icon-only text, on-screen duplicate text). Their source questions are
     hard-blacklisted (AUDIT_EXCLUSIONS below) and replacements are regenerated.
  B. Systematic fixes:
     1. Selection among valid candidates is seeded-RANDOM (rng.choice among the
        best-scoring options), not first-valid.
     2. New filters: (a) source bbox wider than 60% of screen width rejected
        (full-row/field bboxes); (b) UI-state answers {on,off,yes,no,true,false} and
        single-character answers rejected; (c) duplicate-text exclusion is
        unconditional — the same normalized text at 2+ distinct on-screen locations
        (center delta > 2% of W or H, anywhere on screen, even same half) skips the
        candidate; (d) midline +/-10% (half) and +/-5% grid-boundary bands kept.
     3. Stratified balance: half template targets an even top/bottom split
        (N//2 vs N-N//2); grid3x3 fills underrepresented classes first. If balance
        is unreachable with the base 25 screens, the pool expands (same eligibility
        rules, >=3 content questions) up to MAX_SCREENS=35.
     4. Position rule described as ACTUALLY implemented: position is judged from the
        MEDIAN OVER ANNOTATOR BBOX CENTERS (median of per-annotator center-x and
        center-y). The recorded source_bbox is the element-wise median of annotator
        bboxes and is NOT what position is judged from.

Inputs (relative to --root):
  data/screenqa_probe/official/answers_and_bboxes/validation.json
  data/screenqa_probe/official/short_answers/validation.json

Images: any selected screen missing from data/screenqa_pilot/ is downloaded from the
HF community mirror bevaya/RICO-ScreenQA at pinned revision IMAGES_REVISION via
row-group-targeted parquet reads (only the needed row groups' image column chunks are
fetched). Downloaded images are verified against annotation width/height.

Inputs also include the committed, exhaustive post-generation visual audit:
  experiments/manifests/t4_visual_audit.jsonl

Outputs (written only with --output-dir or --write-in-place):
  experiments/manifests/t4_pilot.jsonl        one row per screen (content + location
                                              questions; content questions carry
                                              type_draft per the frozen T4 rule)
  experiments/manifests/t4_pilot.meta.json    rules, stats, achieved balance,
                                              majority baselines, audit exclusions,
                                              statistical plan, image provenance
  experiments/manifests/t4_pairs_draft.jsonl  explicit content->location pair manifest
  experiments/manifests/t4_review.xlsx        human review sheet + persisted
                                              sanity_check sheet (10 seeded recomputes)

CLI:
  python vlm_diagnosis/scripts/gen_t4_pilot.py            # safe preview; writes nothing
  python vlm_diagnosis/scripts/gen_t4_pilot.py --output-dir /tmp/t4-preview
  python vlm_diagnosis/scripts/gen_t4_pilot.py --write-in-place  # explicit overwrite

Requires: pillow, openpyxl, pyarrow, huggingface_hub (network only if images missing).
"""
import argparse
import hashlib
import json
import os
import random
import statistics
import sys
from collections import Counter, defaultdict

# ---------------------------------------------------------------- constants
SEED_DEFAULT = 42
N_BASE = 25           # base screen count (v1 size)
MAX_SCREENS = 35      # balance-driven expansion cap (audit B3)
NO_ANSWER = '<no answer>'
STATE_WORDS = {'on', 'off', 'yes', 'no', 'true', 'false'}   # audit B2(b)
WIDE_BBOX_FRAC = 0.60                                       # audit B2(a)
DUP_DISTINCT_FRAC = 0.02   # centers differing by >2% of W or H = distinct location
MIDLINE_BAND = (0.40, 0.60)   # half template: exclude center-y in 40-60% of H
GRID_BOUNDARY_BAND = 0.05     # grid template: exclude center within 5% of 1/3, 2/3
STATUS_BAR_FRAC = 0.03
MIN_CONTENT_Q = 3
BOOTSTRAP_REPLICATES = 10_000
BOOTSTRAP_SEED = 20260815
EVIDENCE_IOU_SAME = 0.80
EVIDENCE_CONTAIN_SAME = 0.90

ANNOTATIONS_REVISION = '1dfdbccaf56948821b5fa8ffe5d186fe4751e46d'
FULL_ANNOTATIONS_SHA256 = '1e32e5e06ea9bfe3421baceae800ee22de2481663b818ff47f9c6350b36ca138'
SHORT_ANNOTATIONS_SHA256 = '4420ce951f6bad386e2549680f6c31025073501ef0f82a72afaefc43432da743'
VISUAL_AUDIT_SHA256 = '95960d6d39bfc12686dce86ec2da938775f88bf09b8743579df2b9e741edfec3'
IMAGES_REPO = 'bevaya/RICO-ScreenQA'
IMAGES_REVISION = '44c8508ea528b4d4e035cc06e28d7baf8db09ec6'
IMAGES_SHARDS = ['validation-00000-of-00003.parquet',
                 'validation-00001-of-00003.parquet',
                 'validation-00002-of-00003.parquet']

# Source questions of the 9 location questions that failed the 2026-08-15 visual
# audit. Blacklisted as location-question sources (several failure modes — icon-only
# text, duplicates outside the annotated elements — are not machine-detectable from
# the annotations alone).
AUDIT_EXCLUSIONS = {
    'sqa_val_06679': "v1 _loc2 '09:00': full-row bbox spanning the screen -> wrong grid cell",
    'sqa_val_01472': "v1 _loc1 '18': spinner/full-element bbox, not the text extent",
    'sqa_val_05301': "v1 _loc2 '18': spinner/full-element bbox, not the text extent",
    'sqa_val_02356': "v1 _loc2 'Brush': element is an icon; text not visibly rendered",
    'sqa_val_03384': "v1 _loc2 'on': UI toggle state, not visible text; multiple toggles on",
    'sqa_val_05761': "v1 _loc1 'on': UI toggle state, not visible text; multiple toggles on",
    'sqa_val_04404': "v1 _loc1 'Facebook': text appears at top AND bottom of screen (duplicate outside annotated elements)",
    'sqa_val_08120': "v1 _loc1 'Applesauce': text duplicated on screen (duplicate outside annotated elements)",
}
BAD_V1_LOC_IDS = {
    'sqa_val_06679_loc2', 'sqa_val_01472_loc1', 'sqa_val_05301_loc2',
    'sqa_val_02356_loc2', 'sqa_val_03384_loc2', 'sqa_val_05761_loc1',
    'sqa_val_04404_loc1', 'sqa_val_08120_loc1', 'sqa_val_08120_loc2',
}
# note: sqa_val_08120_loc2 ('40 MINS') was not itself flagged, but blacklisting is per
# source question; only the 9 flagged ids are recorded as audit failures in meta.
AUDIT_FAILED_LOC_IDS = sorted(BAD_V1_LOC_IDS - {'sqa_val_08120_loc2'})

# Exclusions from the v2 build's own visual sanity check (2026-08-15): failure modes
# not detectable from the annotations alone, found by inspecting the 10-question
# seeded sample with bboxes drawn on the images.
V2_VISUAL_EXCLUSIONS = {
    'sqa_val_05364': "v2 candidate 'Sat, 11 Feb': same timestamp text rendered on "
                     "two news items (nytimes + washingtonpost rows); duplicate "
                     "outside annotated elements -> ambiguous location",
    'sqa_val_03018': "v2 candidate 'SAM App': text repeats in the headline 'SAM App "
                     "for Anxiety - Free Download' (duplicate outside annotated "
                     "elements; unconditional duplicate rule applies even though "
                     "both occurrences are in the top half)",
    'sqa_val_06821': "v2 candidate '1975': bbox sits on a wallpaper region with no "
                     "visibly rendered text (annotated answer not legible on screen)",
    'sqa_val_06233': "v2 candidate 'Facebook': header 'Log in With Facebook' (top) "
                     "AND footer 'This doesn't let the app post to Facebook.' "
                     "(bottom) — duplicate outside annotated elements, half answer "
                     "ambiguous",
    'sqa_val_06328': "v2 candidate 'Weather': word rendered many times on the "
                     "iTunes preview page (screen title, app names, category) — "
                     "duplicates outside annotated elements",
    'sqa_val_03457': "v2 candidate 'Greg Laurie': text is both the app-bar title "
                     "(top) and the player artist line (bottom) — duplicate "
                     "outside annotated elements, half answer ambiguous",
    'sqa_val_07727': "v2 candidate '1 mg': bbox sits on an empty region of the "
                     "medication list — dose text not visibly rendered",
    'sqa_val_02540': "v2 candidate 'Return': 'Return date' field label elsewhere "
                     "on screen repeats the word (duplicate outside annotated "
                     "elements, grid answer ambiguous)",
    'sqa_val_05020': "v2 candidate 'Grace Chan': profile name (middle-left) AND "
                     "post author line (bottom-left) — duplicate outside annotated "
                     "elements, grid answer ambiguous",
    'sqa_val_01041': "v2 candidate 'LISTEN': tab (top) AND 'Listen Later' item "
                     "(bottom) — whole-word duplicate outside annotated elements",
    'sqa_val_01410': "v2 candidate '30 °F': main temperature AND 'Low: 30 °F' line "
                     "— duplicate outside annotated elements",
    'sqa_val_04347': "v2 candidate '$0.00': invoice shows Total $0.00 AND Balance "
                     "Due $0.00 — duplicate (same cell, but unconditional rule)",
    'sqa_val_04720': "v2 candidate '13:00': '20 May at 13:00' AND 'Yesterday at "
                     "13:00' — duplicate outside annotated elements",
    'sqa_val_04778': "v2 candidate 'Today': tab (top) AND ad text 'Talk to your "
                     "doctor today' (bottom) — whole-word duplicate",
    'sqa_val_05441': "v2 candidate '$6.00': '$6.00 Value' rendered on three deal "
                     "cards — duplicates outside annotated elements",
    'sqa_val_06957': "v2 candidate 'Restaurant.com': title bar, install banner "
                     "(twice) and 'Specials by Restaurant.com' — duplicates in "
                     "both halves",
    'sqa_val_07958': "v2 candidate '2 hours ago': CNN item AND New York Times item "
                     "— duplicate outside annotated elements, grid ambiguous",
    'sqa_val_06467': "v2 candidate 'settings': target is the gear ICON in the "
                     "toolbar — annotated text not visibly rendered",
    'sqa_val_07953': "v2 candidate 'New York Times': byline on two news items "
                     "(top-left and middle) — duplicate outside annotated elements",
    'sqa_val_08138': "v2 candidate 'Wed, Feb 15': date rendered on four article "
                     "cards in both halves — duplicates outside annotated elements",
    'sqa_val_08438': "v2 candidate 'OsmAnd (online tiles)': row title AND 'Tile "
                     "data:' subtitle repeat the exact text — duplicate outside "
                     "annotated elements",
}

# Login-flow brand names: 'Log in with <brand>' headers pair with '... post to
# <brand>' footer disclaimers on the same dialog, so the brand text reliably
# appears in both halves. Failed visual audit three times (v1 sqa_val_04404, v2
# sqa_val_06233 and sqa_val_02822) -> systematic filter.
LOGIN_BRANDS = {'facebook', 'google', 'twitter', 'instagram'}
BLACKLIST = {**AUDIT_EXCLUSIONS, **V2_VISUAL_EXCLUSIONS}

# Screen-level exclusions found during the v2 build (image-vs-annotation checks).
SCREEN_EXCLUSIONS = {
    54145: 'mirror image is landscape 1920x1080 but the official annotation says '
           '1080x1920 — dimension mismatch would corrupt position labels',
}

# Star-rating answers ('5 stars', '4.5 stars', ...) are rendered as star WIDGETS,
# not visible text, and rating widgets typically repeat down list screens — the same
# failure mode as the v1 'Brush' icon case (systematic filter added after the v2
# visual check caught '5 stars').
import re
RATING_RE = re.compile(r'^\d+([.,]\d+)?\s*(stars?|★+)$')

GRID_NAMES = [['top-left', 'top-center', 'top-right'],
              ['middle-left', 'center', 'middle-right'],
              ['bottom-left', 'bottom-center', 'bottom-right']]
GRID_LABELS = [n for row in GRID_NAMES for n in row]
HALF_LABELS = ['top half', 'bottom half']

POSITIONAL_ANSWERS = set(
    HALF_LABELS + GRID_LABELS +
    ['top', 'bottom', 'left', 'right', 'center', 'middle', 'upper', 'lower',
     'left half', 'right half', 'upper half', 'lower half',
     'top left', 'top center', 'top right', 'middle left', 'middle right',
     'bottom left', 'bottom center', 'bottom right',
     'at the top', 'at the bottom', 'on the left', 'on the right'])


def norm(t):
    return ' '.join(t.lower().split())


def content_type_draft(question, answers):
    """Return a conservative primary-type draft plus its uncertainty.

    The frozen rule is about the ability needed to answer, not just answer syntax.
    Selection/status/tab questions therefore remain grounding drafts and are never
    auto-promoted to T4.  OCR and semantic are intentionally kept as one left-side
    group because distinguishing them reliably still requires human review.
    """
    q = norm(question)
    if any(norm(a) in POSITIONAL_ANSWERS for a in answers):
        return {'type_draft': 'layout', 'type_uncertain': False,
                'type_reason': 'answer itself is a position/layout label'}
    if 'how many' in q:
        return {'type_draft': 'count', 'type_uncertain': False,
                'type_reason': "question contains 'how many'"}
    grounding_cue = re.search(
        r'\b(selected|chosen|status|setting|activated|checked|highlighted|tab)\b', q)
    if grounding_cue:
        return {'type_draft': 'grounding', 'type_uncertain': True,
                'type_reason': 'selection/status/tab cue requires visual state grounding'}
    return {'type_draft': 'OCR/semantic', 'type_uncertain': True,
            'type_reason': 'automatic rule cannot reliably separate OCR from semantic'}


def _qid_index(qid):
    return int(qid.split('_')[-1])


def _bbox_iou_and_containment(a, b):
    if not a or not b:
        return 0.0, 0.0
    x1, y1 = max(a[0], b[0]), max(a[1], b[1])
    x2, y2 = min(a[2], b[2]), min(a[3], b[3])
    inter = max(0, x2 - x1) * max(0, y2 - y1)
    area_a = max(0, a[2] - a[0]) * max(0, a[3] - a[1])
    area_b = max(0, b[2] - b[0]) * max(0, b[3] - b[1])
    union = area_a + area_b - inter
    smaller = min(area_a, area_b)
    return (inter / union if union else 0.0,
            inter / smaller if smaller else 0.0)


def _whole_word_related(a, b):
    """Equal text or one normalized string occurring as a whole-word phrase."""
    if not a or not b:
        return False
    if a == b:
        return True
    short, long = sorted((a, b), key=len)
    return bool(re.search(r'(?<!\w)' + re.escape(short) + r'(?!\w)', long))


def canonical_evidence(record):
    """Conservative one-block draft from official per-annotator evidence.

    A canonical block is emitted only when a normalized element text has strict
    majority support among the annotators' single-element selections.
    Raw blocks retain per-annotator support internally so a non-canonical exact
    overlap can be classified as ``partial`` rather than silently ``different``.
    """
    ground_truth = record.get('ground_truth') or []
    single = []
    raw_block_support = Counter()
    multi_element_annotations = 0
    for g in ground_truth:
        ui = g.get('ui_elements') or []
        if len(ui) > 1:
            multi_element_annotations += 1
        # Count annotators, not duplicate elements within one annotation.
        for block in {(norm(e.get('text', '')), tuple(e['bounds'])) for e in ui}:
            raw_block_support[block] += 1
        if len(ui) == 1:
            single.append(ui[0])
    result = {'text': None, 'bbox': None, 'support': 0,
              'annotators': len(ground_truth),
              'raw_block_support': raw_block_support,
              'multi_element_annotations': multi_element_annotations}
    if not single:
        return result
    # Cluster by both normalized text and spatially compatible bboxes. Text-only
    # majority would create a phantom median between two repeated elements.
    clusters = []
    for e in single:
        text = norm(e.get('text', ''))
        bbox = list(e['bounds'])
        compatible = []
        for cluster in clusters:
            if cluster['text'] != text:
                continue
            center = [statistics.median(b[k] for b in cluster['bboxes'])
                      for k in range(4)]
            iou, containment = _bbox_iou_and_containment(bbox, center)
            if iou >= EVIDENCE_IOU_SAME or containment >= EVIDENCE_CONTAIN_SAME:
                compatible.append((iou, containment, cluster))
        if compatible:
            max(compatible, key=lambda item: (item[0], item[1]))[2]['bboxes'].append(bbox)
        else:
            clusters.append({'text': text, 'bboxes': [bbox]})
    best = min(clusters, key=lambda c: (
        -len(c['bboxes']), c['text'], tuple(c['bboxes'][0])))
    best_count = len(best['bboxes'])
    if best_count < len(ground_truth) // 2 + 1:
        return result
    result.update({
        'text': best['text'],
        'bbox': [round(statistics.median(b[k] for b in best['bboxes']))
                 for k in range(4)],
        'support': best_count,
    })
    return result


def compare_evidence(qA_id, qB_source_id, full):
    """Draft same/partial/different relation between official evidence blocks."""
    a = canonical_evidence(full[_qid_index(qA_id)])
    b = canonical_evidence(full[_qid_index(qB_source_id)])
    if qA_id == qB_source_id:
        overlap = 'same'
        reason = 'same official source question id'
    else:
        iou, containment = _bbox_iou_and_containment(a['bbox'], b['bbox'])
        text_related = _whole_word_related(a['text'], b['text'])
        if (a['bbox'] and b['bbox'] and text_related and
                (iou >= EVIDENCE_IOU_SAME or containment >= EVIDENCE_CONTAIN_SAME)):
            overlap = 'same'
            reason = (f'canonical block match: text_related=true, IoU={iou:.3f}, '
                      f'smaller_containment={containment:.3f}')
        elif set(a['raw_block_support']) & set(b['raw_block_support']):
            overlap = 'partial'
            shared = set(a['raw_block_support']) & set(b['raw_block_support'])
            block = max(shared, key=lambda x: (
                min(a['raw_block_support'][x], b['raw_block_support'][x]), x))
            reason = (
                'exact raw block overlap without a canonical whole-block match: '
                f'qA_support={a["raw_block_support"][block]}/{a["annotators"]}, '
                f'qB_support={b["raw_block_support"][block]}/{b["annotators"]}, '
                f'qA_multi_element_annotations={a["multi_element_annotations"]}, '
                f'qB_multi_element_annotations={b["multi_element_annotations"]}; '
                'human review required')
        else:
            overlap = 'different'
            reason = (f'no conservative canonical match: text_related={text_related}, '
                      f'IoU={iou:.3f}, smaller_containment={containment:.3f}')
    internal = {'raw_block_support', 'multi_element_annotations'}
    public_a = {k: v for k, v in a.items() if k not in internal}
    public_b = {k: v for k, v in b.items() if k not in internal}
    return overlap, reason, public_a, public_b


# ---------------------------------------------------------------- candidate rules
def loc_candidate(q, all_elems, excl):
    """Validate one question as a location-question source. Returns (cand, reason).

    Position is judged from the median over annotator bbox CENTERS (median of the
    per-annotator center-x values and center-y values). The recorded source_bbox is
    the element-wise median of annotator bboxes (display/reference only).
    """
    W, H = q['W'], q['H']
    if q['qid'] in BLACKLIST:
        return None, 'audit_blacklist'
    elems = []
    for g in q['gt']:
        ui = g.get('ui_elements') or []
        if len(ui) == 0:
            return None, 'no_bbox'
        if len(ui) > 1:
            return None, 'multi_element_answer'
        elems.append(ui[0])
    if not elems:
        return None, 'no_bbox'
    texts = {norm(e['text']) for e in elems}
    if len(texts) != 1:
        return None, 'annotator_text_mismatch'
    text = elems[0]['text'].strip()
    ntext = norm(text)
    if len(ntext) < 2 or sum(c.isalnum() for c in ntext) < 2:
        # <2 alphanumeric chars: single-char answers and punctuation-padded
        # single glyphs like '-2', which tend to be tiny/illegible on screen
        return None, 'text_too_short_or_single_char'
    if ntext in STATE_WORDS:
        return None, 'ui_state_answer'
    if ntext in LOGIN_BRANDS:
        return None, 'login_brand_answer'
    if RATING_RE.match(ntext):
        return None, 'rating_widget_answer'
    if ntext.isdigit() and len(ntext) <= 2:
        # bare 1-2 digit integers repeat pervasively in UIs (counts, scores,
        # table cells) and duplicates are rarely all annotated — v1 '18' and the
        # v2 visual check's '14' (score table) both failed this way
        return None, 'short_bare_integer'
    bbox = [round(statistics.median(e['bounds'][k] for e in elems)) for k in range(4)]
    if (bbox[2] - bbox[0]) > WIDE_BBOX_FRAC * W:
        return None, 'wide_bbox'
    cxs = [(e['bounds'][0] + e['bounds'][2]) / 2 for e in elems]
    cys = [(e['bounds'][1] + e['bounds'][3]) / 2 for e in elems]
    cx, cy = statistics.median(cxs), statistics.median(cys)
    if cy < STATUS_BAR_FRAC * H:
        return None, 'status_bar'
    # unconditional duplicate-text exclusion (audit B2(c)): same normalized text at
    # 2+ distinct locations anywhere on screen -> skip (threshold only separates
    # annotator jitter on the SAME element from genuinely distinct elements)
    centers = [c for t, c in all_elems if t == ntext]
    for i in range(len(centers)):
        for j in range(i + 1, len(centers)):
            (x1, y1), (x2, y2) = centers[i], centers[j]
            if abs(x1 - x2) > DUP_DISTINCT_FRAC * W or abs(y1 - y2) > DUP_DISTINCT_FRAC * H:
                return None, 'duplicate_text_on_screen'
    # substring ambiguity (v2 visual check round 2, 'Girl' vs 'Panda Girl'): if the
    # candidate text also occurs as a whole-word substring inside a DIFFERENT
    # annotated element's text elsewhere on screen, the quoted text matches multiple
    # visible strings -> ambiguous
    pat = re.compile(r'\b' + re.escape(ntext) + r'\b')
    for t, (ox, oy) in all_elems:
        if t != ntext and pat.search(t) and (
                abs(ox - cx) > DUP_DISTINCT_FRAC * W or
                abs(oy - cy) > DUP_DISTINCT_FRAC * H):
            return None, 'substring_of_other_text'
    cand = {'text': text, 'bbox': bbox, 'cx': cx, 'cy': cy}
    fy, fx = cy / H, cx / W
    # half template
    if MIDLINE_BAND[0] <= fy <= MIDLINE_BAND[1]:
        excl['half_midline_band'] += 1
    elif len({y / H < 0.5 for y in cys}) > 1:
        excl['half_annotator_disagree'] += 1
    else:
        cand['half'] = 'top half' if fy < 0.5 else 'bottom half'
    # grid3x3 template
    def cell(f):
        return 0 if f < 1 / 3 else (1 if f < 2 / 3 else 2)
    if any(abs(f - b) <= GRID_BOUNDARY_BAND for f in (fx, fy) for b in (1 / 3, 2 / 3)):
        excl['grid_boundary_band'] += 1
    elif len({(cell(x / W), cell(y / H)) for x, y in zip(cxs, cys)}) > 1:
        excl['grid_annotator_disagree'] += 1
    else:
        cand['grid'] = GRID_NAMES[cell(fy)][cell(fx)]
    if 'half' not in cand and 'grid' not in cand:
        return None, 'no_valid_template'
    return cand, None


# ---------------------------------------------------------------- balancing
def assign(screens, seed):
    """Greedy stratified assignment of one (half, grid) question pair per screen.

    Screens are processed most-constrained-first (fewest distinct grid labels).
    For each screen the pair whose grid label (then half label) currently has the
    LOWEST count is chosen — fills underrepresented classes first. Ties are broken
    by seeded rng.choice (audit B1: seeded-random, never first-valid).
    """
    rng = random.Random(seed)
    ch, cg = Counter(), Counter()
    order = sorted(
        screens,
        key=lambda e: (len({e['cands'][g][1]['grid'] for _, g in e['options']}),
                       len(e['options']), e['image_id']))
    chosen = {}
    for e in order:
        best, best_key = [], None
        for h, g in e['options']:
            key = (cg[e['cands'][g][1]['grid']], ch[e['cands'][h][1]['half']])
            if best_key is None or key < best_key:
                best_key, best = key, [(h, g)]
            elif key == best_key:
                best.append((h, g))
        h, g = rng.choice(sorted(best))
        chosen[e['image_id']] = (h, g)
        ch[e['cands'][h][1]['half']] += 1
        cg[e['cands'][g][1]['grid']] += 1
    return chosen, ch, cg


def balanced_selection(eligible, seed):
    """Base N_BASE screens (seeded shuffle, 1080x1920 preferred) + balance-driven
    expansion up to MAX_SCREENS. Returns (selected, chosen, ch, cg, expansion_log)."""
    rng = random.Random(seed)
    hi = sorted([e for e in eligible if (e['W'], e['H']) == (1080, 1920)],
                key=lambda e: e['image_id'])
    lo = sorted([e for e in eligible if (e['W'], e['H']) != (1080, 1920)],
                key=lambda e: e['image_id'])
    rng.shuffle(hi)
    rng.shuffle(lo)
    shuffled = hi + lo
    selected, rest = list(shuffled[:N_BASE]), list(shuffled[N_BASE:])
    log = []
    while True:
        chosen, ch, cg = assign(selected, seed + 1)
        half_diff = abs(ch['top half'] - ch['bottom half'])
        min_grid = min(cg[l] for l in GRID_LABELS)
        if (min_grid >= 2 and half_diff <= 1) or len(selected) >= MAX_SCREENS or not rest:
            break
        pick, why = None, None
        for lab in sorted(GRID_LABELS, key=lambda l: cg[l]):
            if cg[lab] >= 2:
                break
            for e in rest:
                if any(e['cands'][g][1]['grid'] == lab for _, g in e['options']):
                    pick, why = e, f'grid:{lab} count={cg[lab]}'
                    break
            if pick:
                break
        if pick is None and half_diff > 1:
            need = 'top half' if ch['top half'] < ch['bottom half'] else 'bottom half'
            for e in rest:
                if any(e['cands'][h][1]['half'] == need for h, _ in e['options']):
                    pick, why = e, f'half:{need}'
                    break
        if pick is None:
            break  # no remaining screen can supply a needed class
        selected.append(pick)
        rest.remove(pick)
        log.append({'added_image_id': pick['image_id'], 'reason': why})
    return selected, chosen, ch, cg, log


# ---------------------------------------------------------------- images
def sha256_file(path):
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1 << 20), b''):
            h.update(chunk)
    return h.hexdigest()


def ensure_images(selected, root, old_meta):
    """Ensure every selected screen's image exists locally; download missing ones
    from the pinned HF mirror revision. Returns provenance dict."""
    img_dir = os.path.join(root, 'data', 'screenqa_pilot')
    os.makedirs(img_dir, exist_ok=True)
    old_prov = (old_meta or {}).get('image_provenance', {})
    if old_prov:
        recorded_revision = (old_meta or {}).get('source', {}).get('images_revision')
        if recorded_revision != IMAGES_REVISION:
            raise RuntimeError(
                'cannot trust recorded image hashes from a different or missing '
                f'mirror revision: {recorded_revision!r} != {IMAGES_REVISION!r}')
    prov, missing, unverified = {}, [], set()
    for e in selected:
        sid = str(e['image_id'])
        path = os.path.join(img_dir, f'{sid}.jpg')
        if os.path.exists(path):
            sha = sha256_file(path)
            old_entry = old_prov.get(sid, {})
            p = dict(old_entry or {'file_name': f'images/rico/{sid}.jpg'})
            p['sha256'] = sha
            expected_sha = old_entry.get('sha256')
            if expected_sha not in (None, sha):
                raise RuntimeError(
                    f'image {sid}: local sha256 {sha} differs from recorded pinned '
                    f'mirror sha256 {expected_sha}')
            if expected_sha is None:
                unverified.add(sid)
            prov[sid] = p
        else:
            missing.append(sid)
    need_scan = (bool(missing) or bool(unverified) or
                 any('shard' not in prov[s] for s in prov))
    if not need_scan:
        return prov
    import pyarrow.parquet as pq
    from huggingface_hub import HfFileSystem
    fs = HfFileSystem()
    base = f'datasets/{IMAGES_REPO}@{IMAGES_REVISION}/data'
    wanted_sids = (set(missing) | unverified |
                   {s for s in prov if 'shard' not in prov[s]})
    wanted = {f'images/rico/{sid}.jpg': sid for sid in wanted_sids}
    print(f'[images] scanning shards for {len(wanted)} screens '
          f'({len(missing)} to download, {len(unverified)} to authenticate)...',
          flush=True)
    for shard in IMAGES_SHARDS:
        if not wanted:
            break
        with fs.open(f'{base}/{shard}', 'rb') as f:
            pf = pq.ParquetFile(f)
            for rg in range(pf.num_row_groups):
                if not wanted:
                    break
                names = pf.read_row_group(rg, columns=['file_name'])
                names = names.column('file_name').to_pylist()
                first = {}  # one row per QA pair -> same image repeats; keep first
                for i, n in enumerate(names):
                    if n in wanted and n not in first:
                        first[n] = i
                if not first:
                    continue
                hits = [(i, n) for n, i in first.items()]
                need_bytes = [(i, n) for i, n in hits
                              if wanted[n] in missing or wanted[n] in unverified]
                tbl = (pf.read_row_group(rg, columns=['file_name', 'image'])
                       if need_bytes else None)
                for i, n in hits:
                    sid = wanted.pop(n)
                    entry = prov.get(sid, {'file_name': n})
                    entry.update({'file_name': n, 'shard': shard, 'row_group': rg})
                    if sid in missing or sid in unverified:
                        img = tbl.column('image')[i].as_py()
                        data = img['bytes'] if isinstance(img, dict) else img
                        mirror_sha = hashlib.sha256(data).hexdigest()
                    if sid in missing:
                        path = os.path.join(img_dir, f'{sid}.jpg')
                        with open(path, 'wb') as out:
                            out.write(data)
                        entry['sha256'] = mirror_sha
                        print(f'[images] downloaded {sid}.jpg '
                              f'({shard} rg{rg}, {len(data)} bytes)', flush=True)
                    elif sid in unverified:
                        if entry['sha256'] != mirror_sha:
                            raise RuntimeError(
                                f'image {sid}: local sha256 {entry["sha256"]} differs '
                                f'from pinned mirror sha256 {mirror_sha}')
                        entry['verified_against_pinned_mirror'] = True
                    prov[sid] = entry
    if wanted:
        raise RuntimeError(f'images not found in mirror: {sorted(wanted.values())}')
    return prov


def validate_existing_images(selected, root, old_meta):
    """Read-only validation for safe preview mode.

    Missing files are reported (an explicit write run may fetch them), while every
    image that already exists must match both its recorded hash, when available,
    and the official annotation dimensions.
    """
    from PIL import Image
    old_prov = (old_meta or {}).get('image_provenance', {})
    if old_prov:
        recorded_revision = (old_meta or {}).get('source', {}).get('images_revision')
        if recorded_revision != IMAGES_REVISION:
            raise RuntimeError(
                'cannot trust recorded image hashes from a different or missing '
                f'mirror revision: {recorded_revision!r} != {IMAGES_REVISION!r}')
    missing = []
    for e in selected:
        sid = str(e['image_id'])
        path = os.path.join(root, 'data', 'screenqa_pilot', f'{sid}.jpg')
        if not os.path.exists(path):
            missing.append(sid)
            continue
        actual_sha = sha256_file(path)
        expected_sha = old_prov.get(sid, {}).get('sha256')
        if expected_sha is not None and actual_sha != expected_sha:
            raise RuntimeError(
                f'image {sid}: local sha256 {actual_sha} differs from recorded '
                f'pinned mirror sha256 {expected_sha}')
        with Image.open(path) as im:
            im.load()
            if im.size != (e['W'], e['H']):
                raise RuntimeError(
                    f'image {sid}: file {im.size} != annotation ({e["W"]}, {e["H"]})')
    unverified = [str(e['image_id']) for e in selected
                  if (os.path.exists(os.path.join(
                      root, 'data', 'screenqa_pilot', f'{e["image_id"]}.jpg')) and
                      old_prov.get(str(e['image_id']), {}).get('sha256') is None)]
    return missing, unverified


def verify_images(selected, root):
    from PIL import Image
    for e in selected:
        path = os.path.join(root, 'data', 'screenqa_pilot', f'{e["image_id"]}.jpg')
        with Image.open(path) as im:
            im.load()
            if im.size != (e['W'], e['H']):
                raise RuntimeError(
                    f'image {e["image_id"]}: file {im.size} != annotation '
                    f'({e["W"]}, {e["H"]})')


def apply_visual_audit(rows, audit_path):
    """Exact-join the committed exhaustive audit and remove failed locations.

    The audit is deliberately post-generation: ``Announce`` must be generated in
    the original seeded assignment and then removed, otherwise preselection changes
    RNG tie-breaking and silently substitutes unaudited questions.
    """
    actual_sha = sha256_file(audit_path)
    if actual_sha != VISUAL_AUDIT_SHA256:
        raise RuntimeError(
            f'visual audit sha256 mismatch: {actual_sha} != {VISUAL_AUDIT_SHA256}')
    records = []
    with open(audit_path) as f:
        for lineno, line in enumerate(f, 1):
            if not line.strip():
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise RuntimeError(f'{audit_path}:{lineno}: invalid JSON: {exc}') from exc

    generated = {}
    for row in rows:
        for loc in row['location_questions']:
            qid = loc['question_id']
            if qid in generated:
                raise RuntimeError(f'duplicate generated location id: {qid}')
            generated[qid] = {
                'question_id': qid,
                'sample_id': row['sample_id'],
                'image': row['image'],
                'template': loc['template'],
                'claimed_answer': loc['answers'][0],
                'source_answer_text': loc['source_answer_text'],
                'source_bbox': loc['source_bbox'],
            }
    audited = {}
    for rec in records:
        qid = rec.get('question_id')
        if not qid:
            raise RuntimeError('visual audit record missing question_id')
        if qid in audited:
            raise RuntimeError(f'duplicate visual audit question_id: {qid}')
        audited[qid] = rec
    missing = sorted(set(generated) - set(audited))
    unknown = sorted(set(audited) - set(generated))
    if missing or unknown:
        raise RuntimeError(
            f'visual audit exact join failed: missing={missing}, unknown={unknown}')

    join_fields = ('question_id', 'sample_id', 'image', 'template',
                   'claimed_answer', 'source_answer_text', 'source_bbox')
    for qid, expected in generated.items():
        rec = audited[qid]
        mismatches = {k: (rec.get(k), expected[k]) for k in join_fields
                      if rec.get(k) != expected[k]}
        if mismatches:
            raise RuntimeError(f'visual audit payload mismatch for {qid}: {mismatches}')
        verdict = rec.get('visual_verdict')
        if verdict not in {'pass', 'fail'}:
            raise RuntimeError(f'visual audit unknown verdict for {qid}: {verdict!r}')
        if not rec.get('verified_by') or not rec.get('verified_date'):
            raise RuntimeError(f'visual audit missing verifier/date for {qid}')

    kept_rows = []
    for row in rows:
        row['location_questions'] = [
            loc for loc in row['location_questions']
            if audited[loc['question_id']]['visual_verdict'] == 'pass'
        ]
        if row['location_questions']:
            kept_rows.append(row)
    rows[:] = kept_rows
    all_loc = [(row, loc) for row in rows for loc in row['location_questions']]
    counts = Counter(rec['visual_verdict'] for rec in records)
    return all_loc, records, counts, actual_sha


def majority_baseline(counts, labels):
    """Constant-majority raw accuracy, balanced accuracy and fixed-label macro-F1."""
    total = sum(counts.values())
    if not total:
        raise RuntimeError('cannot compute a majority baseline on an empty template')
    predicted = max(labels, key=lambda label: counts[label])
    prevalence = counts[predicted] / total
    majority_f1 = 2 * prevalence / (1 + prevalence)
    return {
        'predicted_label': predicted,
        'raw_accuracy': round(prevalence, 4),
        'balanced_accuracy': round(1 / len(labels), 4),
        'macro_f1': round(majority_f1 / len(labels), 4),
        'fixed_label_universe': list(labels),
    }


# ---------------------------------------------------------------- main build
def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    default_root = os.path.abspath(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..'))
    ap.add_argument('--root', default=default_root)
    ap.add_argument('--seed', type=int, default=SEED_DEFAULT)
    mode = ap.add_mutually_exclusive_group()
    mode.add_argument('--output-dir',
                      help='write generated manifest files to this directory')
    mode.add_argument('--write-in-place', action='store_true',
                      help='explicitly overwrite --root/experiments/manifests outputs')
    mode.add_argument('--dry-run', action='store_true',
                      help='validate and preview only (also the default with no mode)')
    ap.add_argument('--visual-audit',
                    help='audit JSONL input (default: committed manifest directory)')
    args = ap.parse_args()
    root, seed = os.path.abspath(args.root), args.seed
    write_outputs = bool(args.output_dir or args.write_in_place)
    if args.output_dir:
        expanded_output = os.path.expanduser(args.output_dir)
        out_dir = (expanded_output if os.path.isabs(expanded_output) else
                   os.path.join(root, expanded_output))
        out_dir = os.path.abspath(out_dir)
    else:
        out_dir = os.path.join(root, 'experiments', 'manifests')
    audit_path = os.path.abspath(
        args.visual_audit or
        os.path.join(root, 'experiments', 'manifests', 't4_visual_audit.jsonl'))
    if not write_outputs:
        print('[safe preview] no files will be written; use --output-dir or '
              '--write-in-place to persist outputs')

    full_path = os.path.join(
        root, 'data', 'screenqa_probe', 'official', 'answers_and_bboxes',
        'validation.json')
    short_path = os.path.join(
        root, 'data', 'screenqa_probe', 'official', 'short_answers',
        'validation.json')
    source_hashes = {
        'answers_and_bboxes_validation_sha256': sha256_file(full_path),
        'short_answers_validation_sha256': sha256_file(short_path),
    }
    expected_hashes = {
        'answers_and_bboxes_validation_sha256': FULL_ANNOTATIONS_SHA256,
        'short_answers_validation_sha256': SHORT_ANNOTATIONS_SHA256,
    }
    if source_hashes != expected_hashes:
        raise RuntimeError(
            f'annotation input sha256 mismatch: actual={source_hashes}, '
            f'expected={expected_hashes}')
    with open(full_path) as f:
        full = json.load(f)
    with open(short_path) as f:
        short = json.load(f)
    if len(full) != len(short):
        raise RuntimeError(f'full/short length mismatch: {len(full)} != {len(short)}')
    for a, b in zip(full, short):
        if a['image_id'] != b['image_id'] or a['question'] != b['question']:
            raise RuntimeError('full/short annotation alignment mismatch')

    screens = defaultdict(list)
    for idx, (f, s) in enumerate(zip(full, short)):
        screens[f['image_id']].append({
            'qid': f'sqa_val_{idx:05d}',
            'question': f['question'],
            'short_answers': [g for g in s['ground_truth']
                              if g.strip() and g.strip() != NO_ANSWER],
            'gt': f['ground_truth'],
            'W': f['image_width'], 'H': f['image_height'],
        })

    excl, screen_excl = Counter(), Counter()
    eligible = []
    for image_id, qs in sorted(screens.items()):
        if image_id in SCREEN_EXCLUSIONS:
            screen_excl['screen_blacklist'] += 1
            continue
        W, H = qs[0]['W'], qs[0]['H']
        content_qs = [q for q in qs if q['short_answers']]
        if len(content_qs) < MIN_CONTENT_Q:
            screen_excl['lt3_content_questions'] += 1
            continue
        all_elems = []
        for q in qs:
            for g in q['gt']:
                for e in (g.get('ui_elements') or []):
                    all_elems.append((norm(e['text']),
                                      ((e['bounds'][0] + e['bounds'][2]) / 2,
                                       (e['bounds'][1] + e['bounds'][3]) / 2)))
        cands = {}
        for q in qs:
            c, reason = loc_candidate(q, all_elems, excl)
            if c is None:
                excl[reason] += 1
            else:
                cands[q['qid']] = (q, c)
        halfq = sorted(q for q, (_, c) in cands.items() if 'half' in c)
        gridq = sorted(q for q, (_, c) in cands.items() if 'grid' in c)
        options = [(h, g) for h in halfq for g in gridq if h != g]
        if not options:
            screen_excl['no_valid_half_grid_pair'] += 1
            continue
        eligible.append({'image_id': image_id, 'W': W, 'H': H,
                         'content_qs': content_qs, 'cands': cands,
                         'options': options})

    print(f'screens scanned: {len(screens)}, eligible: {len(eligible)} '
          f'(1080x1920: {sum(1 for e in eligible if (e["W"], e["H"]) == (1080, 1920))})')

    selected, chosen, ch, cg, expansion_log = balanced_selection(eligible, seed)
    print(f'selected screens: {len(selected)} (expanded by {len(expansion_log)})')
    print('half:', dict(ch), ' grid:', dict(cg))

    # ---- build per-screen records (still pre-audit: exactly 2 locations/screen)
    rows, all_loc, all_content = [], [], []
    for e in sorted(selected, key=lambda e: e['image_id']):
        hqid, gqid = chosen[e['image_id']]
        loc_recs = []
        for qid, tmpl in ((hqid, 'half'), (gqid, 'grid3x3')):
            q, c = e['cands'][qid]
            if tmpl == 'half':
                question = (f"Is the '{c['text']}' shown in the top half or the "
                            f"bottom half of the screen?")
                ans = c['half']
            else:
                question = (f"In which region of the 3x3 grid of the screen is "
                            f"'{c['text']}' shown: top-left, top-center, top-right, "
                            f"middle-left, center, middle-right, bottom-left, "
                            f"bottom-center, or bottom-right?")
                ans = c['grid']
            loc_recs.append({
                'question_id': f'{qid}_loc{1 if tmpl == "half" else 2}',
                'question': question, 'answers': [ans],
                'source_question_id': qid, 'source_answer_text': c['text'],
                'source_bbox': c['bbox'], 'template': tmpl})
        content_recs = []
        for q in e['content_qs']:
            answers = list(dict.fromkeys(q['short_answers']))
            type_info = content_type_draft(q['question'], answers)
            content_recs.append({
                'question_id': q['qid'], 'question': q['question'],
                'answers': answers, **type_info})
        row = {
            'dataset': 'ScreenQA', 'dataset_revision': ANNOTATIONS_REVISION,
            'source_split': 'validation', 'split': 'pilot',
            'sample_id': str(e['image_id']),
            'image': f'data/screenqa_pilot/{e["image_id"]}.jpg',
            'image_width': e['W'], 'image_height': e['H'],
            'image_mirror': IMAGES_REPO, 'image_mirror_revision': IMAGES_REVISION,
            'content_questions': content_recs,
            'location_questions': loc_recs,
            'selection_seed': seed,
        }
        rows.append(row)
        all_loc.extend([(row, r) for r in loc_recs])
        all_content.extend([(row, r) for r in content_recs])

    pre_audit_all_loc = list(all_loc)
    pre_half = Counter(r['answers'][0] for _, r in pre_audit_all_loc
                       if r['template'] == 'half')
    pre_grid = Counter(r['answers'][0] for _, r in pre_audit_all_loc
                       if r['template'] == 'grid3x3')
    pre_loc_ids = {r['question_id'] for _, r in pre_audit_all_loc}
    if not pre_loc_ids.isdisjoint(BAD_V1_LOC_IDS):
        raise RuntimeError('a v1 audit-failed location question was regenerated')

    # ---- exact post-generation visual-audit join; failures are removed, not replaced
    all_loc, visual_records, visual_counts, visual_audit_sha = apply_visual_audit(
        rows, audit_path)
    all_content = [(row, r) for row in rows for r in row['content_questions']]
    final_loc_ids = {r['question_id'] for _, r in all_loc}
    failed_loc_ids = {r['question_id'] for r in visual_records
                      if r['visual_verdict'] == 'fail'}
    if final_loc_ids & failed_loc_ids:
        raise RuntimeError('visual-audit failure survived post-generation filtering')

    # ---- pair manifest (content -> location only; evidence-aware draft)
    pairs, n_pending = [], 0
    left_types = {'OCR/semantic', 'count'}
    for row in rows:
        for cq in row['content_questions']:
            for lq in row['location_questions']:
                same_source = (lq['source_question_id'] == cq['question_id'])
                overlap, overlap_reason, ev_a, ev_b = compare_evidence(
                    cq['question_id'], lq['source_question_id'], full)
                type_crossing = cq['type_draft'] in left_types
                pending = overlap == 'same' and type_crossing
                if pending:
                    n_pending += 1
                    draft_label = ''
                    rationale = ('same evidence block plus left->layout type crossing; '
                                 'T2-vs-T4 precedence awaits user decision')
                elif overlap == 'same':
                    draft_label = 'T2'
                    rationale = ('same evidence block without an automatically safe '
                                 'type crossing')
                elif overlap == 'partial':
                    draft_label = 'uncertain'
                    rationale = 'partial evidence overlap requires human adjudication'
                elif type_crossing:
                    draft_label = 'T4'
                    rationale = (f"different evidence, {cq['type_draft']} -> layout "
                                 'type crossing')
                else:
                    draft_label = 'uncertain'
                    rationale = (f"different evidence but {cq['type_draft']} -> layout "
                                 'is not a safe cross-group T4 assignment')
                pairs.append({
                    'dataset': 'ScreenQA', 'dataset_revision': ANNOTATIONS_REVISION,
                    'source_split': 'validation', 'split': 'pilot',
                    'sample_id': row['sample_id'], 'image': row['image'],
                    'pair_id': f"{row['sample_id']}_{cq['question_id']}_{lq['question_id']}",
                    'direction': 'content_to_location',
                    'qA_id': cq['question_id'], 'qA': cq['question'],
                    'qA_answers': ' | '.join(cq['answers']),
                    'qA_type_draft': cq['type_draft'],
                    'qA_type_uncertain': cq['type_uncertain'],
                    'qA_type_reason': cq['type_reason'],
                    'qA_evidence_text_draft': ev_a['text'],
                    'qA_evidence_bbox_draft': ev_a['bbox'],
                    'qA_evidence_support': ev_a['support'],
                    'qA_evidence_annotators': ev_a['annotators'],
                    'qB_id': lq['question_id'], 'qB': lq['question'],
                    'qB_answers': lq['answers'][0], 'qB_type_draft': 'layout',
                    'qB_template': lq['template'],
                    'qB_source_question_id': lq['source_question_id'],
                    'qB_evidence_text_draft': ev_b['text'],
                    'qB_evidence_bbox_draft': ev_b['bbox'],
                    'qB_evidence_support': ev_b['support'],
                    'qB_evidence_annotators': ev_b['annotators'],
                    'evidence_overlap_draft': overlap,
                    'evidence_overlap_reason': overlap_reason,
                    'same_source': same_source,
                    'pending_precedence_decision': pending,
                    'draft_label': draft_label, 'draft_rationale': rationale,
                    'uncertain': True,
                    'label_A': '', 'notes_A': '', 'label_B': '', 'notes_B': '',
                    'adjudicated_label': '', 'adjudication_note': '',
                    'selection_seed': seed,
                })

    final_half = Counter(r['answers'][0] for _, r in all_loc
                         if r['template'] == 'half')
    final_grid = Counter(r['answers'][0] for _, r in all_loc
                         if r['template'] == 'grid3x3')
    print(f'post-audit: {len(all_loc)}/{len(pre_audit_all_loc)} locations retained; '
          f'verdicts={dict(visual_counts)}')
    print('final half:', dict(final_half), ' final grid:', dict(final_grid))
    print(f'final pairs: {len(pairs)} ({n_pending} pending precedence)')

    source_meta_path = os.path.join(
        root, 'experiments', 'manifests', 't4_pilot.meta.json')
    if os.path.exists(source_meta_path):
        with open(source_meta_path) as f:
            old_meta = json.load(f)
    else:
        old_meta = None
    selected_ids = {int(row['sample_id']) for row in rows}
    final_selected = [e for e in selected if e['image_id'] in selected_ids]
    missing_images, unverified_images = validate_existing_images(
        final_selected, root, old_meta)
    if not write_outputs:
        if unverified_images:
            raise RuntimeError(
                'safe preview cannot authenticate existing images without recorded '
                'hashes; use an explicit write mode to compare them with the pinned '
                f'mirror: {unverified_images}')
        if missing_images:
            print(f'[safe preview] {len(missing_images)} selected images are absent; '
                  'an explicit write run is required to fetch them')
        print('[safe preview] validation complete; no files written')
        return 0

    os.makedirs(out_dir, exist_ok=True)
    prov = ensure_images(final_selected, root, old_meta)
    verify_images(final_selected, root)

    pilot_out = os.path.join(out_dir, 't4_pilot.jsonl')
    pairs_out = os.path.join(out_dir, 't4_pairs_draft.jsonl')
    review_out = os.path.join(out_dir, 't4_review.xlsx')
    meta_out = os.path.join(out_dir, 't4_pilot.meta.json')
    with open(pilot_out, 'w') as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + '\n')
    with open(pairs_out, 'w') as f:
        for pair in pairs:
            f.write(json.dumps(pair, ensure_ascii=False) + '\n')

    # ---- review workbook: final questions, arithmetic check, exhaustive visual audit
    import openpyxl
    from PIL import Image
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 't4_location_review'
    ws.append(['sample_id', 'question_id', 'image', 'template', 'question',
               'generated_answer', 'source_answer_text', 'source_bbox',
               'valid_position_correct', 'valid_unambiguous', 'reviewer_notes'])
    for row, r in all_loc:
        ws.append([row['sample_id'], r['question_id'], row['image'], r['template'],
                   r['question'], r['answers'][0], r['source_answer_text'],
                   json.dumps(r['source_bbox']), None, None, None])

    ws2 = wb.create_sheet('sanity_check')
    ws2.append(['question_id', 'sample_id', 'template', 'claimed', 'recomputed',
                'center_fx', 'center_fy', 'match', 'method'])
    method = ('independent recompute: actual image W/H (PIL); median over raw '
              'annotator bbox centers; fixed center-to-label rule; '
              f'seed={seed + 2}')
    sample = random.Random(seed + 2).sample(
        sorted(all_loc, key=lambda t: t[1]['question_id']), min(10, len(all_loc)))
    n_match = 0
    for row, r in sample:
        gt = full[_qid_index(r['source_question_id'])]['ground_truth']
        cxs = [(g['ui_elements'][0]['bounds'][0] +
                g['ui_elements'][0]['bounds'][2]) / 2 for g in gt]
        cys = [(g['ui_elements'][0]['bounds'][1] +
                g['ui_elements'][0]['bounds'][3]) / 2 for g in gt]
        with Image.open(os.path.join(root, row['image'])) as im:
            W, H = im.size
        fx, fy = statistics.median(cxs) / W, statistics.median(cys) / H
        if r['template'] == 'half':
            recomputed = 'top half' if fy < 0.5 else 'bottom half'
        else:
            def cell(f):
                return 0 if f < 1 / 3 else (1 if f < 2 / 3 else 2)
            recomputed = GRID_NAMES[cell(fy)][cell(fx)]
        match = recomputed == r['answers'][0]
        n_match += match
        ws2.append([r['question_id'], row['sample_id'], r['template'],
                    r['answers'][0], recomputed, round(fx, 4), round(fy, 4),
                    'yes' if match else 'NO', method])

    ws3 = wb.create_sheet('visual_audit')
    visual_headers = ['question_id', 'sample_id', 'image', 'template',
                      'claimed_answer', 'source_answer_text', 'source_bbox',
                      'visual_verdict', 'verified_by', 'verified_date', 'note',
                      'retained_after_audit']
    ws3.append(visual_headers)
    for rec in visual_records:
        values = [json.dumps(rec[k], ensure_ascii=False)
                  if isinstance(rec.get(k), (dict, list)) else rec.get(k, '')
                  for k in visual_headers[:-1]]
        ws3.append(values + [rec['question_id'] in final_loc_ids])
    wb.save(review_out)
    print(f'sanity_check: {n_match}/{len(sample)} recomputed labels match')

    # ---- final authoritative metadata; pre-audit figures live in a separate block
    meta = {
        'created': '2026-08-15',
        'version': 2,
        'generator': 'vlm_diagnosis/scripts/gen_t4_pilot.py',
        'selection_seed': seed,
        'source': {
            'annotations': 'google-research-datasets/screen_qa validation full+short',
            'annotations_revision': ANNOTATIONS_REVISION,
            **source_hashes,
            'images': f'HF community mirror {IMAGES_REPO}, validation parquet shards',
            'images_revision': IMAGES_REVISION,
            'residual_risk': 'community mirror rather than official RICO release; '
                             'automatic duplicate checks see annotated evidence only, '
                             'so exhaustive post-generation visual audit is mandatory',
        },
        'visual_audit': {
            'input': (os.path.relpath(audit_path, root)
                      if os.path.commonpath((root, audit_path)) == root else audit_path),
            'sha256': visual_audit_sha,
            'join_policy': 'exact question-id set and exact payload match; missing, '
                           'unknown, duplicate, malformed, or unverified records fail',
            'generated': len(visual_records),
            'passed': visual_counts['pass'],
            'failed': visual_counts['fail'],
            'failed_location_ids': sorted(failed_loc_ids),
            'policy': 'failed locations are removed post-generation and are never '
                      'replaced without a new exhaustive audit',
        },
        'question_id_convention': 'sqa_val_<0-based validation index>; loc1=half, '
                                  'loc2=grid3x3',
        'selection_rule': f'>=3 usable content questions and a distinct-qid half/grid '
                          f'candidate pair; 1080x1920 first; seeded base {N_BASE}; '
                          f'greedy balance expansion capped at {MAX_SCREENS}',
        'location_generation_rules': {
            'position_rule': 'median over per-annotator bbox centers; source_bbox is '
                             'the rounded element-wise median for display only',
            'half_template': 'top/bottom; reject center-y in closed 40-60% band and '
                             'annotator half disagreement',
            'grid3x3_template': 'thirds; reject centers within closed +/-5% boundary '
                                'bands and annotator cell disagreement',
            'duplicate_scope': 'automatic exact/substring checks cover annotated '
                               'evidence elements only; visual audit covers the screen',
            'candidate_selection': 'locally greedy underrepresented-grid then half; '
                                   'seeded random tie break',
        },
        'audit_exclusions': {
            'failed_v1_location_questions': AUDIT_FAILED_LOC_IDS,
            'blacklisted_source_questions': AUDIT_EXCLUSIONS,
            'v2_preselection_visual_exclusions': V2_VISUAL_EXCLUSIONS,
            'screen_exclusions': {str(k): v for k, v in SCREEN_EXCLUSIONS.items()},
            'post_generation_failures': {
                rec['question_id']: rec.get('note', '') for rec in visual_records
                if rec['visual_verdict'] == 'fail'},
        },
        'pairs_manifest': {
            'file': os.path.basename(pairs_out),
            'direction': 'content_to_location only, fixed as the pilot causal order. '
                         'Reverse direction directly leaks qA only for same/partial-'
                         'evidence pairs; cross-evidence reverse pairs are omitted for '
                         'directional consistency, not because all contain qA answers.',
            'evidence_rule': 'same qid => same; otherwise canonical text plus IoU>=0.8 '
                             'or smaller-box containment>=0.9 => same; annotator-only '
                             'exact intersection => partial; otherwise different',
            'pending_precedence_decision': 'only same-evidence plus left-group '
                                           '(OCR/semantic/count)->layout crossings',
            'type_policy': 'selection/status/tab cues are grounding+uncertain and '
                           'cannot receive automatic T4; every pair remains uncertain '
                           'until two-reviewer adjudication',
        },
        'statistical_plan': {
            'unit_of_analysis': 'screen-macro; locations and content-location pairs '
                                'within a screen are not independent',
            'metric_definition': {
                'raw_accuracy': 'mean of each screen\'s mean correctness; each screen '
                                'has total weight 1 regardless of question count',
                'balanced_accuracy': 'macro recall on a screen-weighted confusion '
                                     'matrix; each question weight is 1 divided by '
                                     'the number of evaluated questions on its screen',
                'macro_f1': 'fixed-label macro-F1 on the same screen-weighted '
                            'confusion matrix',
                'per_template': 'also report half and grid3x3 separately; within each '
                                'template each retained screen contributes at most '
                                'one question',
                'zero_support': 'fixed labels with zero gold or prediction support '
                                'remain in the average with recall/F1=0 '
                                '(zero_division=0)',
            },
            'bootstrap': {
                'replicates': BOOTSTRAP_REPLICATES,
                'seed': BOOTSTRAP_SEED,
                'confidence_level': 0.95,
                'interval': 'percentile [2.5%,97.5%]',
                'resampling_unit': 'screen with replacement',
                'labels': 'fixed post-audit label universe and gold labels in every '
                          'replicate; no relabeling or class dropping',
                'paired_delta': 'for method comparisons, reuse the identical sampled '
                                'screen indices per replicate and report method-A '
                                'minus method-B',
            },
            'metrics': ['raw_accuracy', 'balanced_accuracy', 'macro_f1'],
        },
        'pre_audit_generation_stats': {
            'screens_selected': len(selected),
            'location_questions': len(pre_audit_all_loc),
            'half_distribution': {l: pre_half.get(l, 0) for l in HALF_LABELS},
            'grid_distribution': {l: pre_grid.get(l, 0) for l in GRID_LABELS},
            'expansion_beyond_base': expansion_log,
        },
        'stats': {
            'authoritative_stage': 'post_visual_audit',
            'screens_scanned': len(screens),
            'screens_eligible': len(eligible),
            'eligible_1080x1920': sum(
                1 for e in eligible if (e['W'], e['H']) == (1080, 1920)),
            'screens_selected': len(rows),
            'screen_location_count_distribution': dict(Counter(
                len(row['location_questions']) for row in rows)),
            'content_questions': len(all_content),
            'location_questions': len(all_loc),
            'content_type_draft_distribution': dict(Counter(
                r['type_draft'] for _, r in all_content)),
            'content_type_uncertain': sum(
                1 for _, r in all_content if r['type_uncertain']),
            'pairs_total': len(pairs),
            'pairs_pending_precedence': n_pending,
            'pair_draft_label_distribution': dict(Counter(
                p['draft_label'] or 'pending' for p in pairs)),
            'evidence_overlap_draft_distribution': dict(Counter(
                p['evidence_overlap_draft'] for p in pairs)),
            'achieved_balance': {
                'half': {
                    'distribution': {l: final_half.get(l, 0) for l in HALF_LABELS},
                    'majority_baseline': majority_baseline(final_half, HALF_LABELS),
                },
                'grid3x3': {
                    'distribution': {l: final_grid.get(l, 0) for l in GRID_LABELS},
                    'majority_baseline': majority_baseline(final_grid, GRID_LABELS),
                },
            },
            'screen_exclusions': dict(screen_excl),
            'question_level_location_exclusions': dict(excl),
            'sanity_check': {'sampled': len(sample), 'matched': n_match,
                             'sheet': 't4_review.xlsx#sanity_check'},
        },
        'image_provenance': {k: prov[k] for k in sorted(prov, key=int)},
    }
    with open(meta_out, 'w') as f:
        json.dump(meta, f, indent=1, ensure_ascii=False)

    print(f'wrote {len(rows)} screens, {len(all_loc)} audited location questions, '
          f'{len(all_content)} content questions, {len(pairs)} pairs to {out_dir}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
