"""Build independent, blind A/B review workbooks for the T4 pair manifest.

The source of truth is ``experiments/manifests/t4_pairs_draft.jsonl``.  The
generated workbooks deliberately omit draft labels/rationales, the other
reviewer's fields, and adjudication fields.  Each reviewer receives the same
pair context and an initially blank, reviewer-owned set of type/evidence/label
columns.

Examples:

    python -m vlm_diagnosis.scripts.build_t4_pair_reviews
    python vlm_diagnosis/scripts/build_t4_pair_reviews.py \
        --root /root/research/heewon/VLM \
        --output-dir experiments/manifests
"""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_RELATIVE_PATH = Path("experiments/manifests/t4_pairs_draft.jsonl")
OUTPUT_NAMES = {
    "A": "t4_pairs_review_A.xlsx",
    "B": "t4_pairs_review_B.xlsx",
}

# Only factual, reviewer-independent context is copied from the draft manifest.
# In particular, type drafts and same-source/precedence hints are omitted because
# they would anchor the independent label decision.
COMMON_COLUMNS = [
    "dataset",
    "dataset_revision",
    "source_split",
    "split",
    "pair_id",
    "sample_id",
    "image",
    "direction",
    "qA_id",
    "qA",
    "qA_answers",
    "qB_id",
    "qB",
    "qB_answers",
    "qB_template",
    "selection_seed",
]

REVIEW_COLUMNS = [
    "qA_primary_type",
    "qB_primary_type",
    "qA_evidence_block",
    "qB_evidence_block",
    "evidence_overlap",
    "final_label",
    "notes",
]

# These fields must never be exposed in a blind review workbook.  Keeping the
# assertion close to workbook construction prevents a future source-schema
# expansion from silently leaking draft or adjudication information.
FORBIDDEN_COLUMNS = {
    "qA_type_draft",
    "qB_type_draft",
    "same_source",
    "pending_precedence_decision",
    "draft_label",
    "draft_rationale",
    "uncertain",
    "label_A",
    "notes_A",
    "label_B",
    "notes_B",
    "adjudicated_label",
    "adjudication_note",
}

TYPE_VALUES = ["OCR", "semantic", "layout", "grounding", "icon", "count"]
OVERLAP_VALUES = ["same", "partial", "different"]
LABEL_VALUES = ["T0", "T1", "T2", "T3", "T4"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")

COLUMN_WIDTHS = {
    "dataset": 12,
    "dataset_revision": 18,
    "source_split": 12,
    "split": 10,
    "pair_id": 38,
    "sample_id": 12,
    "image": 38,
    "direction": 20,
    "qA_id": 20,
    "qA": 52,
    "qA_answers": 48,
    "qB_id": 24,
    "qB": 72,
    "qB_answers": 22,
    "qB_template": 13,
    "selection_seed": 14,
    "qA_primary_type": 19,
    "qB_primary_type": 19,
    "qA_evidence_block": 34,
    "qB_evidence_block": 34,
    "evidence_overlap": 19,
    "final_label": 14,
    "notes": 48,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build blind A/B T4 pair-review workbooks."
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=DEFAULT_ROOT,
        help="Repository root (default: inferred from this script).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("experiments/manifests"),
        help="Output directory; relative paths are resolved under --root.",
    )
    return parser.parse_args()


def load_pairs(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"T4 pair manifest not found: {path}")

    rows: list[dict[str, Any]] = []
    seen_pair_ids: set[str] = set()
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSON at {path}:{line_number}: {exc}") from exc

            missing = [column for column in COMMON_COLUMNS if column not in row]
            if missing:
                raise ValueError(
                    f"Missing required fields at {path}:{line_number}: "
                    + ", ".join(missing)
                )

            pair_id = str(row["pair_id"])
            if pair_id in seen_pair_ids:
                raise ValueError(f"Duplicate pair_id at {path}:{line_number}: {pair_id}")
            seen_pair_ids.add(pair_id)
            rows.append(row)

    if not rows:
        raise ValueError(f"T4 pair manifest is empty: {path}")
    return rows


def manifest_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def force_text(cell: Any, value: Any) -> None:
    """Write source text as literal text, including strings beginning with '='."""
    if value is None:
        cell.value = ""
    elif isinstance(value, str):
        cell.value = value
        cell.data_type = "s"
    else:
        cell.value = value


def add_list_validation(
    worksheet: Any,
    column_name: str,
    values: Iterable[str],
    row_count: int,
    prompt: str,
) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}
    column_index = headers[column_name]
    column_letter = get_column_letter(column_index)
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        errorStyle="stop",
        errorTitle="허용되지 않은 값",
        error="드롭다운의 동결된 값만 선택하세요.",
        promptTitle=column_name,
        prompt=prompt,
        showErrorMessage=True,
        showInputMessage=True,
    )
    worksheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{row_count + 1}")


def add_instructions_sheet(
    workbook: Workbook,
    reviewer: str,
    source_path: Path,
    source_sha256: str,
    pair_count: int,
) -> None:
    worksheet = workbook.create_sheet("instructions")
    worksheet.append(["항목", "내용"])
    rows = [
        ("reviewer", reviewer),
        ("source_manifest", str(source_path)),
        ("source_sha256", source_sha256),
        ("pair_count", pair_count),
        (
            "blind_review",
            "다른 검수자의 파일과 master의 draft/adjudication 값을 보지 않고 독립 판정",
        ),
        ("type_vocab", "OCR / semantic / layout / grounding / icon / count 중 primary 1개"),
        ("evidence_unit", "문단 / 표 셀 / 제목·머리글 / 범례 / 그림·UI element 등 의미 블록"),
        ("overlap", "same / partial / different; partial이면 notes에 불확실성 기록"),
        ("T0", "질문이 동일함"),
        ("T1", "같은 답·같은 evidence이며 wording만 바뀐 paraphrase"),
        ("T2", "paraphrase가 아닌 다른 질문이지만 evidence block이 same"),
        ("T3", "evidence block이 different이고 T4 유형 교차가 아님"),
        (
            "T4",
            "primary type이 {OCR, semantic, count}와 {layout, grounding, icon} 사이를 교차",
        ),
        (
            "precedence_pending",
            "same evidence이면서 type도 교차하면 T2↔T4 우선순위가 아직 미결이다. "
            "두 type·evidence·overlap은 입력하되 final_label은 비워 두고 notes에 "
            "PRECEDENCE_PENDING을 기록한다.",
        ),
        (
            "workflow",
            "두 evidence block과 primary type을 먼저 적고 overlap, final_label, notes 순서로 입력",
        ),
    ]
    for item, content in rows:
        worksheet.append([item, content])

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in worksheet.iter_rows(min_row=2):
        row[0].fill = SECTION_FILL
        row[0].font = Font(bold=True)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 100
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:B{worksheet.max_row}"


def build_workbook(
    rows: list[dict[str, Any]],
    reviewer: str,
    source_path: Path,
    source_sha256: str,
) -> Workbook:
    headers = COMMON_COLUMNS + REVIEW_COLUMNS
    leaked = set(headers) & FORBIDDEN_COLUMNS
    if leaked:
        raise AssertionError(f"Blind-review schema leaked forbidden fields: {sorted(leaked)}")

    workbook = Workbook()
    workbook.properties.creator = "VLM T4 pair review builder"
    workbook.properties.title = f"T4 blind pair review — reviewer {reviewer}"
    workbook.properties.subject = "Independent T0-T4 pair labeling"
    # Fixed document properties keep workbook metadata stable across runs.
    fixed_time = datetime(2026, 8, 15, 0, 0, 0)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time

    worksheet = workbook.active
    worksheet.title = "pairs"
    worksheet.append(headers)

    for row_index, source_row in enumerate(rows, start=2):
        for column_index, column_name in enumerate(COMMON_COLUMNS, start=1):
            force_text(
                worksheet.cell(row=row_index, column=column_index),
                source_row[column_name],
            )
        for column_index in range(len(COMMON_COLUMNS) + 1, len(headers) + 1):
            worksheet.cell(row=row_index, column=column_index, value="")

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    input_start = len(COMMON_COLUMNS) + 1
    for row in worksheet.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in row[input_start - 1 :]:
            cell.fill = INPUT_FILL

    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS[header]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    worksheet.sheet_view.zoomScale = 80
    worksheet.row_dimensions[1].height = 34

    add_list_validation(
        worksheet,
        "qA_primary_type",
        TYPE_VALUES,
        len(rows),
        "qA 답을 내는 데 필요한 주된 정보 유형을 선택하세요.",
    )
    add_list_validation(
        worksheet,
        "qB_primary_type",
        TYPE_VALUES,
        len(rows),
        "qB 답을 내는 데 필요한 주된 정보 유형을 선택하세요.",
    )
    add_list_validation(
        worksheet,
        "evidence_overlap",
        OVERLAP_VALUES,
        len(rows),
        "same, partial, different 중 하나를 선택하세요.",
    )
    add_list_validation(
        worksheet,
        "final_label",
        LABEL_VALUES,
        len(rows),
        "T0, T1, T2, T3, T4 중 하나를 선택하세요.",
    )

    add_instructions_sheet(workbook, reviewer, source_path, source_sha256, len(rows))
    return workbook


def main() -> int:
    args = parse_args()
    root = args.root.expanduser().resolve()
    output_dir = args.output_dir.expanduser()
    if not output_dir.is_absolute():
        output_dir = root / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    source_path = root / SOURCE_RELATIVE_PATH
    rows = load_pairs(source_path)
    source_digest = manifest_sha256(source_path)

    written: list[Path] = []
    for reviewer, output_name in OUTPUT_NAMES.items():
        workbook = build_workbook(rows, reviewer, SOURCE_RELATIVE_PATH, source_digest)
        output_path = output_dir / output_name
        workbook.save(output_path)
        written.append(output_path)

    print(f"source: {source_path}")
    print(f"pairs: {len(rows)}")
    for output_path in written:
        print(f"wrote: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
