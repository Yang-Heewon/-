"""Validate the committed ScreenQA T4 pilot artifacts against one another.

The validator treats the final manifests as the source of truth for derived counts.
It does not regenerate the sample or hard-code the current pilot size.  In particular,
post-generation visual exclusions must be reflected consistently in the pilot, pair
manifest, metadata, and review workbook.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable


EXPECTED_VERSION = 2
PILOT_REL = Path("experiments/manifests/t4_pilot.jsonl")
PAIRS_REL = Path("experiments/manifests/t4_pairs_draft.jsonl")
AUDIT_REL = Path("experiments/manifests/t4_visual_audit.jsonl")
META_REL = Path("experiments/manifests/t4_pilot.meta.json")
REVIEW_REL = Path("experiments/manifests/t4_review.xlsx")
PAIR_REVIEW_RELS = {
    "A": Path("experiments/manifests/t4_pairs_review_A.xlsx"),
    "B": Path("experiments/manifests/t4_pairs_review_B.xlsx"),
}

HALF_LABELS = ("top half", "bottom half")
GRID_LABELS = (
    "top-left", "top-center", "top-right",
    "middle-left", "center", "middle-right",
    "bottom-left", "bottom-center", "bottom-right",
)
LEFT_CONTENT_TYPES = frozenset({"OCR/semantic", "count"})
RIGHT_CONTENT_TYPES = frozenset({"grounding", "layout", "icon"})
CONTENT_TYPES = LEFT_CONTENT_TYPES | RIGHT_CONTENT_TYPES
EVIDENCE_OVERLAPS = frozenset({"same", "partial", "different"})

AUDIT_FIELDS = (
    "question_id", "sample_id", "image", "template", "claimed_answer",
    "source_answer_text", "source_bbox", "visual_verdict", "verified_by",
    "verified_date", "note",
)
AUDIT_SHEET_FIELDS = AUDIT_FIELDS + ("retained_after_audit",)

PAIR_REVIEW_COMMON_COLUMNS = (
    "dataset", "dataset_revision", "source_split", "split", "pair_id",
    "sample_id", "image", "direction", "qA_id", "qA", "qA_answers",
    "qB_id", "qB", "qB_answers", "qB_template", "selection_seed",
)
PAIR_REVIEW_INPUT_COLUMNS = (
    "qA_primary_type", "qB_primary_type", "qA_evidence_block",
    "qB_evidence_block", "evidence_overlap", "final_label", "notes",
)
PAIR_REVIEW_COLUMNS = PAIR_REVIEW_COMMON_COLUMNS + PAIR_REVIEW_INPUT_COLUMNS
PAIR_REVIEW_PRIVATE_COLUMNS = frozenset({
    "qA_type_draft", "qA_type_uncertain", "qA_type_reason",
    "qA_evidence_text_draft", "qA_evidence_bbox_draft",
    "qA_evidence_support", "qA_evidence_annotators", "qB_type_draft",
    "qB_source_question_id", "qB_evidence_text_draft",
    "qB_evidence_bbox_draft", "qB_evidence_support",
    "qB_evidence_annotators", "evidence_overlap_draft",
    "evidence_overlap_reason", "same_source", "pending_precedence_decision",
    "draft_label", "draft_rationale", "uncertain", "label_A", "notes_A",
    "label_B", "notes_B", "adjudicated_label", "adjudication_note",
})


class Validation:
    def __init__(self) -> None:
        self.errors: list[str] = []

    def check(self, condition: bool, message: str) -> None:
        if not condition:
            self.errors.append(message)

    def equal(self, actual: Any, expected: Any, label: str) -> None:
        if actual != expected:
            self.errors.append(f"{label}: got {actual!r}, expected {expected!r}")

    def close(self, actual: Any, expected: float, label: str, *, tol: float = 5e-4) -> None:
        if not isinstance(actual, (int, float)) or not math.isclose(
                float(actual), expected, rel_tol=0.0, abs_tol=tol):
            self.errors.append(f"{label}: got {actual!r}, expected approximately {expected!r}")


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, 1):
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{path}:{line_number}: invalid JSON: {exc}") from exc
            if not isinstance(row, dict):
                raise ValueError(f"{path}:{line_number}: expected a JSON object")
            rows.append(row)
    if not rows:
        raise ValueError(f"{path}: no records")
    return rows


def load_object(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        value = json.load(handle)
    if not isinstance(value, dict):
        raise ValueError(f"{path}: expected a JSON object")
    return value


def duplicates(values: Iterable[Any]) -> list[Any]:
    counts = Counter(values)
    return sorted((value for value, count in counts.items() if count > 1), key=repr)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def normalized_cell(value: Any) -> Any:
    """Treat an empty JSON string and an empty spreadsheet cell equivalently."""
    return "" if value is None else value


def single_answer(question: dict[str, Any]) -> Any:
    answers = question.get("answers")
    return answers[0] if isinstance(answers, list) and len(answers) == 1 else None


def validate_evidence_fields(v: Validation, pair: dict[str, Any], prefix: str,
                             label: str) -> None:
    text_value = pair.get(f"{prefix}_evidence_text_draft")
    bbox_value = pair.get(f"{prefix}_evidence_bbox_draft")
    support = pair.get(f"{prefix}_evidence_support")
    annotators = pair.get(f"{prefix}_evidence_annotators")
    v.check(text_value is None or isinstance(text_value, str),
            f"{label}.{prefix}_evidence_text_draft must be a string or null")
    bbox_valid = bbox_value is None or (
        isinstance(bbox_value, list) and len(bbox_value) == 4 and
        all(isinstance(value, (int, float)) for value in bbox_value)
    )
    v.check(bbox_valid,
            f"{label}.{prefix}_evidence_bbox_draft must be four numbers or null")
    v.check(isinstance(support, int) and not isinstance(support, bool) and support >= 0,
            f"{label}.{prefix}_evidence_support must be a non-negative integer")
    v.check(isinstance(annotators, int) and not isinstance(annotators, bool)
            and annotators > 0,
            f"{label}.{prefix}_evidence_annotators must be a positive integer")
    if isinstance(support, int) and isinstance(annotators, int):
        v.check(support <= annotators,
                f"{label}.{prefix}_evidence_support exceeds annotators")


def validate_majority_baseline(v: Validation, obj: Any, counts: Counter[str],
                               labels: tuple[str, ...], label: str) -> None:
    if not isinstance(obj, dict):
        v.errors.append(f"{label} is missing or invalid")
        return
    require_keys(
        v,
        obj,
        ("predicted_label", "raw_accuracy", "balanced_accuracy", "macro_f1",
         "fixed_label_universe"),
        label,
    )
    total = sum(counts.values())
    if total <= 0:
        v.errors.append(f"{label}: cannot derive a baseline from zero questions")
        return
    predicted = max(labels, key=lambda value: counts[value])
    prevalence = counts[predicted] / total
    majority_f1 = 2 * prevalence / (1 + prevalence)
    v.equal(obj.get("predicted_label"), predicted, f"{label}.predicted_label")
    v.close(obj.get("raw_accuracy"), prevalence, f"{label}.raw_accuracy")
    v.close(obj.get("balanced_accuracy"), 1 / len(labels),
            f"{label}.balanced_accuracy")
    v.close(obj.get("macro_f1"), majority_f1 / len(labels), f"{label}.macro_f1")
    v.equal(obj.get("fixed_label_universe"), list(labels),
            f"{label}.fixed_label_universe")


def require_keys(v: Validation, obj: dict[str, Any], keys: Iterable[str], label: str) -> None:
    missing = sorted(set(keys) - set(obj))
    v.check(not missing, f"{label}: missing keys {missing}")


def validate(root: Path) -> tuple[Validation, dict[str, int]]:
    v = Validation()
    pilot = load_jsonl(root / PILOT_REL)
    pairs = load_jsonl(root / PAIRS_REL)
    audit = load_jsonl(root / AUDIT_REL)
    meta = load_object(root / META_REL)

    v.equal(meta.get("version"), EXPECTED_VERSION, "meta.version")
    require_keys(
        v,
        meta,
        ("selection_seed", "stats", "pre_audit_generation_stats",
         "image_provenance", "visual_audit", "pairs_manifest"),
        "meta",
    )

    # ---------------------------------------------------------------- pilot
    sample_ids = [str(row.get("sample_id")) for row in pilot]
    images = [row.get("image") for row in pilot]
    v.check(not duplicates(sample_ids), f"pilot: duplicate sample_id values {duplicates(sample_ids)}")
    v.check(not duplicates(images), f"pilot: duplicate image paths {duplicates(images)}")

    pilot_by_sample: dict[str, dict[str, Any]] = {}
    content_by_key: dict[tuple[str, str], tuple[dict[str, Any], dict[str, Any]]] = {}
    location_by_key: dict[tuple[str, str], tuple[dict[str, Any], dict[str, Any]]] = {}
    content_ids: list[str] = []
    location_ids: list[str] = []
    content_types: Counter[str] = Counter()
    half_answers: Counter[str] = Counter()
    grid_answers: Counter[str] = Counter()

    pilot_required = (
        "dataset", "dataset_revision", "source_split", "split", "sample_id", "image",
        "image_width", "image_height", "content_questions", "location_questions",
        "selection_seed",
    )
    for row_number, row in enumerate(pilot, 1):
        label = f"pilot row {row_number}"
        require_keys(v, row, pilot_required, label)
        sample_id = str(row.get("sample_id"))
        pilot_by_sample[sample_id] = row
        v.equal(row.get("selection_seed"), meta.get("selection_seed"), f"{label}.selection_seed")
        v.check(isinstance(row.get("content_questions"), list), f"{label}.content_questions is not a list")
        v.check(isinstance(row.get("location_questions"), list), f"{label}.location_questions is not a list")
        if not isinstance(row.get("content_questions"), list) or not isinstance(
                row.get("location_questions"), list):
            continue

        row_content_ids: set[str] = set()
        for question in row["content_questions"]:
            require_keys(
                v,
                question,
                ("question_id", "question", "answers", "type_draft",
                 "type_uncertain", "type_reason"),
                f"{label} content question",
            )
            qid = str(question.get("question_id"))
            row_content_ids.add(qid)
            content_ids.append(qid)
            content_types[str(question.get("type_draft"))] += 1
            v.check(question.get("type_draft") in CONTENT_TYPES,
                    f"{label} content {qid}: unknown type_draft "
                    f"{question.get('type_draft')!r}")
            v.check(isinstance(question.get("type_uncertain"), bool),
                    f"{label} content {qid}: type_uncertain is not boolean")
            v.check(isinstance(question.get("type_reason"), str)
                    and bool(question.get("type_reason")),
                    f"{label} content {qid}: type_reason is empty or invalid")
            content_by_key[(sample_id, qid)] = (row, question)

        row_location_ids: list[str] = []
        row_sources: list[str] = []
        for question in row["location_questions"]:
            require_keys(
                v,
                question,
                ("question_id", "question", "answers", "source_question_id",
                 "source_answer_text", "source_bbox", "template"),
                f"{label} location question",
            )
            qid = str(question.get("question_id"))
            source_id = str(question.get("source_question_id"))
            row_location_ids.append(qid)
            row_sources.append(source_id)
            location_ids.append(qid)
            location_by_key[(sample_id, qid)] = (row, question)
            v.check(source_id in row_content_ids,
                    f"{label}: location {qid} source {source_id} is not a content question")
            answers = question.get("answers")
            v.check(isinstance(answers, list) and len(answers) == 1,
                    f"{label}: location {qid} must have exactly one answer")
            if isinstance(answers, list) and len(answers) == 1:
                if question.get("template") == "half":
                    half_answers[str(answers[0])] += 1
                elif question.get("template") == "grid3x3":
                    grid_answers[str(answers[0])] += 1
                else:
                    v.errors.append(f"{label}: location {qid} has unknown template "
                                    f"{question.get('template')!r}")
        v.check(not duplicates(row_location_ids),
                f"{label}: duplicate location question ids {duplicates(row_location_ids)}")
        v.check(not duplicates(row_sources),
                f"{label}: location questions reuse source ids {duplicates(row_sources)}")

    v.check(not duplicates(content_ids), f"pilot: duplicate content question ids {duplicates(content_ids)}")
    v.check(not duplicates(location_ids),
            f"pilot: duplicate location question ids {duplicates(location_ids)}")
    v.check(set(content_ids).isdisjoint(location_ids),
            "pilot: content and location question-id sets overlap")

    # Images are content-addressed by the metadata and dimensions are checked against
    # the actual decoded files, not only against manifest fields.
    provenance = meta.get("image_provenance")
    if not isinstance(provenance, dict):
        v.errors.append("meta.image_provenance is not an object")
        provenance = {}
    v.equal(set(provenance), set(sample_ids), "meta.image_provenance sample-id set")
    try:
        from PIL import Image
    except ImportError as exc:  # pragma: no cover - environment failure
        raise RuntimeError("Pillow is required to validate T4 images") from exc
    image_hashes: list[str] = []
    for sample_id, row in pilot_by_sample.items():
        image_value = row.get("image")
        if not isinstance(image_value, str):
            v.errors.append(f"sample {sample_id}: image path is not a string")
            continue
        image_path = root / image_value
        if not image_path.is_file():
            v.errors.append(f"sample {sample_id}: image does not exist: {image_path}")
            continue
        try:
            with Image.open(image_path) as image:
                image.load()
                actual_size = image.size
        except Exception as exc:  # Pillow intentionally reports several exception types
            v.errors.append(f"sample {sample_id}: image decode failed: {exc}")
            continue
        expected_size = (row.get("image_width"), row.get("image_height"))
        v.equal(actual_size, expected_size, f"sample {sample_id}: image dimensions")
        actual_hash = sha256_file(image_path)
        image_hashes.append(actual_hash)
        entry = provenance.get(sample_id)
        if not isinstance(entry, dict):
            v.errors.append(f"sample {sample_id}: missing provenance object")
        else:
            v.equal(entry.get("sha256"), actual_hash, f"sample {sample_id}: provenance sha256")
            if "file_name" in entry:
                v.check(isinstance(entry.get("file_name"), str),
                        f"sample {sample_id}: provenance file_name is not a string")
    v.check(not duplicates(image_hashes), f"pilot: duplicate image hashes {duplicates(image_hashes)}")

    # ---------------------------------------------------------------- pairs
    pair_ids = [str(pair.get("pair_id")) for pair in pairs]
    v.check(not duplicates(pair_ids), f"pairs: duplicate pair ids {duplicates(pair_ids)}")
    expected_triples = {
        (str(row["sample_id"]), str(content["question_id"]), str(location["question_id"]))
        for row in pilot
        for content in row.get("content_questions", [])
        for location in row.get("location_questions", [])
    }
    actual_triples: set[tuple[str, str, str]] = set()
    pending_count = 0
    uncertain_count = 0
    pair_label_distribution: Counter[str] = Counter()
    evidence_overlap_distribution: Counter[str] = Counter()
    direction_counts: Counter[str] = Counter()
    pair_required = (
        "dataset", "dataset_revision", "source_split", "split", "sample_id", "image",
        "pair_id", "direction", "qA_id", "qA", "qA_answers", "qA_type_draft",
        "qA_type_uncertain", "qA_type_reason", "qA_evidence_text_draft",
        "qA_evidence_bbox_draft", "qA_evidence_support", "qA_evidence_annotators",
        "qB_id", "qB", "qB_answers", "qB_type_draft", "qB_template",
        "qB_source_question_id", "qB_evidence_text_draft",
        "qB_evidence_bbox_draft", "qB_evidence_support", "qB_evidence_annotators",
        "evidence_overlap_draft", "evidence_overlap_reason", "same_source",
        "pending_precedence_decision", "draft_label", "draft_rationale", "uncertain",
        "label_A", "notes_A", "label_B", "notes_B", "adjudicated_label",
        "adjudication_note", "selection_seed",
    )
    for row_number, pair in enumerate(pairs, 1):
        label = f"pair row {row_number}"
        require_keys(v, pair, pair_required, label)
        sample_id = str(pair.get("sample_id"))
        q_a_id = str(pair.get("qA_id"))
        q_b_id = str(pair.get("qB_id"))
        triple = (sample_id, q_a_id, q_b_id)
        actual_triples.add(triple)
        direction = str(pair.get("direction"))
        direction_counts[direction] += 1
        v.equal(direction, "content_to_location", f"{label}.direction")
        v.equal(pair.get("selection_seed"), meta.get("selection_seed"),
                f"{label}.selection_seed")

        content_item = content_by_key.get((sample_id, q_a_id))
        location_item = location_by_key.get((sample_id, q_b_id))
        if content_item is None:
            v.errors.append(f"{label}: qA {sample_id}/{q_a_id} is absent from pilot")
            continue
        if location_item is None:
            v.errors.append(f"{label}: qB {sample_id}/{q_b_id} is absent from pilot")
            continue
        pilot_row, content = content_item
        _, location = location_item
        expected_pair_id = f"{sample_id}_{q_a_id}_{q_b_id}"
        v.equal(pair.get("pair_id"), expected_pair_id, f"{label}.pair_id")
        v.equal(pair.get("image"), pilot_row.get("image"), f"{label}.image")
        v.equal(pair.get("qA"), content.get("question"), f"{label}.qA")
        answers = content.get("answers")
        expected_q_a_answers = " | ".join(answers) if isinstance(answers, list) else None
        v.equal(pair.get("qA_answers"), expected_q_a_answers, f"{label}.qA_answers")
        v.equal(pair.get("qA_type_draft"), content.get("type_draft"),
                f"{label}.qA_type_draft")
        v.equal(pair.get("qA_type_uncertain"), content.get("type_uncertain"),
                f"{label}.qA_type_uncertain")
        v.equal(pair.get("qA_type_reason"), content.get("type_reason"),
                f"{label}.qA_type_reason")
        validate_evidence_fields(v, pair, "qA", label)
        v.equal(pair.get("qB"), location.get("question"), f"{label}.qB")
        loc_answers = location.get("answers")
        expected_q_b_answer = loc_answers[0] if isinstance(loc_answers, list) and loc_answers else None
        v.equal(pair.get("qB_answers"), expected_q_b_answer, f"{label}.qB_answers")
        v.equal(pair.get("qB_type_draft"), "layout", f"{label}.qB_type_draft")
        v.equal(pair.get("qB_template"), location.get("template"), f"{label}.qB_template")
        v.equal(pair.get("qB_source_question_id"), location.get("source_question_id"),
                f"{label}.qB_source_question_id")
        validate_evidence_fields(v, pair, "qB", label)

        same_source = location.get("source_question_id") == q_a_id
        v.equal(pair.get("same_source"), same_source, f"{label}.same_source")
        overlap = pair.get("evidence_overlap_draft")
        q_a_type = content.get("type_draft")
        v.check(overlap in EVIDENCE_OVERLAPS,
                f"{label}.evidence_overlap_draft has invalid value {overlap!r}")
        v.check(isinstance(pair.get("evidence_overlap_reason"), str)
                and bool(pair.get("evidence_overlap_reason")),
                f"{label}.evidence_overlap_reason is empty or invalid")
        if same_source:
            v.equal(overlap, "same", f"{label}.evidence_overlap_draft for same source")
            for suffix in ("text_draft", "bbox_draft", "support", "annotators"):
                v.equal(pair.get(f"qA_evidence_{suffix}"),
                        pair.get(f"qB_evidence_{suffix}"),
                        f"{label}.same-source evidence {suffix}")

        # Pending precedence is evidence-aware.  same_source is only a diagnostic
        # identity flag: different source questions may still share one evidence block.
        expected_pending = overlap == "same" and q_a_type in LEFT_CONTENT_TYPES
        v.equal(pair.get("pending_precedence_decision"), expected_pending,
                f"{label}.pending_precedence_decision")
        v.equal(pair.get("uncertain"), True, f"{label}.uncertain")
        pending_count += bool(pair.get("pending_precedence_decision"))
        uncertain_count += bool(pair.get("uncertain"))
        if expected_pending:
            expected_label = ""
        elif overlap == "same":
            expected_label = "T2"
        elif overlap == "partial":
            expected_label = "uncertain"
        elif overlap == "different" and q_a_type in LEFT_CONTENT_TYPES:
            expected_label = "T4"
        else:
            expected_label = "uncertain"
        v.equal(pair.get("draft_label"), expected_label, f"{label}.draft_label")
        v.check(isinstance(pair.get("draft_rationale"), str)
                and bool(pair.get("draft_rationale")),
                f"{label}.draft_rationale is empty or invalid")
        pair_label_distribution[str(pair.get("draft_label") or "pending")] += 1
        evidence_overlap_distribution[str(overlap)] += 1

    missing_pairs = sorted(expected_triples - actual_triples)
    extra_pairs = sorted(actual_triples - expected_triples)
    v.check(not missing_pairs, f"pairs: missing pilot Cartesian rows {missing_pairs[:10]}")
    v.check(not extra_pairs, f"pairs: extra rows outside pilot Cartesian product {extra_pairs[:10]}")
    v.equal(len(actual_triples), len(pairs), "pairs: unique joined triples")
    v.equal(direction_counts, Counter({"content_to_location": len(pairs)}),
            "pairs: direction distribution")

    # ---------------------------------------------------------------- visual audit
    audit_ids = [str(record.get("question_id")) for record in audit]
    v.check(not duplicates(audit_ids), f"visual audit: duplicate question ids {duplicates(audit_ids)}")
    pre_audit_half_answers: Counter[str] = Counter()
    pre_audit_grid_answers: Counter[str] = Counter()
    for row_number, record in enumerate(audit, 1):
        v.equal(set(record), set(AUDIT_FIELDS),
                f"visual audit row {row_number}: field schema")
        require_keys(v, record, AUDIT_FIELDS, f"visual audit row {row_number}")
        v.check(bool(record.get("verified_by")),
                f"visual audit row {row_number}: verified_by is empty")
        v.check(bool(record.get("verified_date")),
                f"visual audit row {row_number}: verified_date is empty")
        template = record.get("template")
        if template == "half":
            pre_audit_half_answers[str(record.get("claimed_answer"))] += 1
            v.check(record.get("claimed_answer") in HALF_LABELS,
                    f"visual audit row {row_number}: invalid half answer "
                    f"{record.get('claimed_answer')!r}")
        elif template == "grid3x3":
            pre_audit_grid_answers[str(record.get("claimed_answer"))] += 1
            v.check(record.get("claimed_answer") in GRID_LABELS,
                    f"visual audit row {row_number}: invalid grid answer "
                    f"{record.get('claimed_answer')!r}")
        else:
            v.errors.append(
                f"visual audit row {row_number}: invalid template {template!r}")
    verdict_counts = Counter(str(record.get("visual_verdict")) for record in audit)
    v.check(set(verdict_counts) <= {"pass", "fail"},
            f"visual audit: unexpected verdicts {sorted(set(verdict_counts) - {'pass', 'fail'})}")
    pass_ids = {str(record.get("question_id")) for record in audit
                if record.get("visual_verdict") == "pass"}
    fail_ids = {str(record.get("question_id")) for record in audit
                if record.get("visual_verdict") == "fail"}
    final_location_ids = set(location_ids)
    v.equal(pass_ids, final_location_ids, "visual-audit pass set vs final location set")
    v.check(fail_ids.isdisjoint(final_location_ids),
            f"visual-audit failed questions retained in final pilot: {sorted(fail_ids & final_location_ids)}")
    v.equal(set(audit_ids), pass_ids | fail_ids,
            "visual audit: every record must have a pass/fail verdict")
    for row_number, record in enumerate(audit, 1):
        if record.get("visual_verdict") != "pass":
            continue
        qid = str(record.get("question_id"))
        sample_id = str(record.get("sample_id"))
        item = location_by_key.get((sample_id, qid))
        if item is None:
            v.errors.append(f"visual audit row {row_number}: passed question {sample_id}/{qid} missing")
            continue
        pilot_row, location = item
        field_pairs = (
            ("image", pilot_row.get("image")),
            ("template", location.get("template")),
            ("claimed_answer", single_answer(location)),
            ("source_answer_text", location.get("source_answer_text")),
            ("source_bbox", location.get("source_bbox")),
        )
        for field, expected in field_pairs:
            v.equal(record.get(field), expected, f"visual audit row {row_number}.{field}")

    # ---------------------------------------------------------------- metadata
    stats = meta.get("stats")
    if not isinstance(stats, dict):
        v.errors.append("meta.stats is not an object")
        stats = {}
    final_stat_keys = (
        "authoritative_stage", "screens_selected", "screen_location_count_distribution",
        "content_questions", "location_questions", "content_type_draft_distribution",
        "content_type_uncertain", "pairs_total", "pairs_pending_precedence",
        "pair_draft_label_distribution", "evidence_overlap_draft_distribution",
        "achieved_balance", "sanity_check",
    )
    require_keys(v, stats, final_stat_keys, "meta.stats")
    v.equal(stats.get("authoritative_stage"), "post_visual_audit",
            "meta.stats.authoritative_stage")
    legacy_stat_keys = {
        "pairs_draft_T4", "pairs_draft_uncertain", "final_location_questions",
        "half_distribution", "grid_distribution", "majority_baseline_half",
        "majority_baseline_grid", "expansion_beyond_base",
    }
    v.check(not (legacy_stat_keys & set(stats)),
            f"meta.stats contains legacy/pre-audit fields "
            f"{sorted(legacy_stat_keys & set(stats))}")
    screen_location_counts = Counter(
        len(row.get("location_questions", [])) for row in pilot)
    derived_stats: dict[str, Any] = {
        "screens_selected": len(pilot),
        "screen_location_count_distribution": {
            str(key): value for key, value in sorted(screen_location_counts.items())
        },
        "content_questions": len(content_ids),
        "location_questions": len(location_ids),
        "content_type_uncertain": sum(
            question.get("type_uncertain") is True
            for row in pilot
            for question in row.get("content_questions", [])
        ),
        "pairs_total": len(pairs),
        "pairs_pending_precedence": pending_count,
        "pair_draft_label_distribution": dict(pair_label_distribution),
        "evidence_overlap_draft_distribution": dict(evidence_overlap_distribution),
    }
    for key, expected in derived_stats.items():
        v.equal(stats.get(key), expected, f"meta.stats.{key}")
    v.equal(uncertain_count, len(pairs), "pairs: all rows uncertain for human review")

    expected_type_distribution = dict(sorted(content_types.items()))
    actual_type_distribution = stats.get("content_type_draft_distribution")
    if isinstance(actual_type_distribution, dict):
        v.equal(dict(sorted(actual_type_distribution.items())), expected_type_distribution,
                "meta.stats.content_type_draft_distribution")
    else:
        v.errors.append("meta.stats.content_type_draft_distribution is missing or invalid")

    half_distribution = {key: half_answers[key] for key in HALF_LABELS}
    grid_distribution = {key: grid_answers[key] for key in GRID_LABELS}

    achieved = stats.get("achieved_balance")
    if isinstance(achieved, dict):
        half_meta = achieved.get("half")
        grid_meta = achieved.get("grid3x3")
        if isinstance(half_meta, dict):
            v.equal(half_meta.get("distribution"), half_distribution,
                    "meta.stats.achieved_balance.half.distribution")
            validate_majority_baseline(
                v, half_meta.get("majority_baseline"), half_answers, HALF_LABELS,
                "meta.stats.achieved_balance.half.majority_baseline",
            )
        else:
            v.errors.append("meta.stats.achieved_balance.half is missing or invalid")
        if isinstance(grid_meta, dict):
            v.equal(grid_meta.get("distribution"), grid_distribution,
                    "meta.stats.achieved_balance.grid3x3.distribution")
            validate_majority_baseline(
                v, grid_meta.get("majority_baseline"), grid_answers, GRID_LABELS,
                "meta.stats.achieved_balance.grid3x3.majority_baseline",
            )
        else:
            v.errors.append("meta.stats.achieved_balance.grid3x3 is missing or invalid")
    else:
        v.errors.append("meta.stats.achieved_balance is missing or invalid")

    pre_audit_stats = meta.get("pre_audit_generation_stats")
    pre_audit_keys = {
        "screens_selected", "location_questions", "half_distribution",
        "grid_distribution", "expansion_beyond_base",
    }
    if not isinstance(pre_audit_stats, dict):
        v.errors.append("meta.pre_audit_generation_stats is missing or invalid")
    else:
        v.equal(set(pre_audit_stats), pre_audit_keys,
                "meta.pre_audit_generation_stats field set")
        v.equal(pre_audit_stats.get("screens_selected"),
                len({str(record.get('sample_id')) for record in audit}),
                "meta.pre_audit_generation_stats.screens_selected")
        v.equal(pre_audit_stats.get("location_questions"), len(audit),
                "meta.pre_audit_generation_stats.location_questions")
        v.equal(pre_audit_stats.get("half_distribution"),
                {key: pre_audit_half_answers[key] for key in HALF_LABELS},
                "meta.pre_audit_generation_stats.half_distribution")
        v.equal(pre_audit_stats.get("grid_distribution"),
                {key: pre_audit_grid_answers[key] for key in GRID_LABELS},
                "meta.pre_audit_generation_stats.grid_distribution")
        v.check(isinstance(pre_audit_stats.get("expansion_beyond_base"), list),
                "meta.pre_audit_generation_stats.expansion_beyond_base is not a list")

    visual_meta = meta.get("visual_audit")
    if not isinstance(visual_meta, dict):
        v.errors.append("meta.visual_audit is not an object")
    else:
        visual_meta_keys = {
            "input", "sha256", "join_policy", "generated", "passed", "failed",
            "failed_location_ids", "policy",
        }
        v.equal(set(visual_meta), visual_meta_keys, "meta.visual_audit field set")
        recorded_audit_input = Path(str(visual_meta.get("input")))
        if recorded_audit_input.is_absolute():
            v.check(recorded_audit_input.is_file(),
                    f"meta.visual_audit.input does not exist: {recorded_audit_input}")
            if recorded_audit_input.is_file():
                v.equal(sha256_file(recorded_audit_input), sha256_file(root / AUDIT_REL),
                        "meta.visual_audit.input content vs validated audit")
        else:
            v.equal(recorded_audit_input, AUDIT_REL, "meta.visual_audit.input")
        v.equal(visual_meta.get("sha256"), sha256_file(root / AUDIT_REL),
                "meta.visual_audit.sha256")
        v.equal(visual_meta.get("generated"), len(audit), "meta.visual_audit.generated")
        v.equal(visual_meta.get("passed"), verdict_counts.get("pass", 0),
                "meta.visual_audit.passed")
        v.equal(visual_meta.get("failed"), verdict_counts.get("fail", 0),
                "meta.visual_audit.failed")
        v.equal(visual_meta.get("failed_location_ids"), sorted(fail_ids),
                "meta.visual_audit.failed_location_ids")

    pairs_meta = meta.get("pairs_manifest")
    if isinstance(pairs_meta, dict):
        v.equal(Path(str(pairs_meta.get("file"))), Path(PAIRS_REL.name),
                "meta.pairs_manifest.file")
        v.check("content_to_location" in str(pairs_meta.get("direction", "")),
                "meta.pairs_manifest.direction does not declare content_to_location")
    else:
        v.errors.append("meta.pairs_manifest is missing or invalid")

    # ---------------------------------------------------------------- workbook
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment failure
        raise RuntimeError("openpyxl is required to validate t4_review.xlsx") from exc
    workbook = load_workbook(root / REVIEW_REL, read_only=True, data_only=True)
    required_sheets = {"t4_location_review", "sanity_check", "visual_audit"}
    v.check(required_sheets <= set(workbook.sheetnames),
            f"review workbook: missing sheets {sorted(required_sheets - set(workbook.sheetnames))}")

    def sheet_records(sheet_name: str, required_headers: Iterable[str],
                      *, exact_headers: Iterable[str] | None = None) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value) if value is not None else "" for value in next(rows)]
        except StopIteration:
            v.errors.append(f"review workbook {sheet_name}: empty sheet")
            return []
        missing_headers = sorted(set(required_headers) - set(headers))
        v.check(not missing_headers,
                f"review workbook {sheet_name}: missing headers {missing_headers}")
        if exact_headers is not None:
            v.equal(headers, list(exact_headers),
                    f"review workbook {sheet_name}: exact header schema")
        return [dict(zip(headers, values)) for values in rows
                if any(value is not None for value in values)]

    location_sheet = sheet_records(
        "t4_location_review",
        ("sample_id", "question_id", "image", "template", "question", "generated_answer",
         "source_answer_text", "source_bbox"),
    )
    v.equal(len(location_sheet), len(location_ids), "review t4_location_review data rows")
    sheet_location_ids = [str(record.get("question_id")) for record in location_sheet]
    v.check(not duplicates(sheet_location_ids),
            f"review t4_location_review: duplicate question ids "
            f"{duplicates(sheet_location_ids)}")
    v.equal(set(sheet_location_ids), final_location_ids,
            "review t4_location_review question-id set")
    for row_number, record in enumerate(location_sheet, 2):
        sample_id = str(record.get("sample_id"))
        qid = str(record.get("question_id"))
        item = location_by_key.get((sample_id, qid))
        if item is None:
            v.errors.append(f"review t4_location_review row {row_number}: unknown {sample_id}/{qid}")
            continue
        pilot_row, location = item
        expected_bbox = json.dumps(location.get("source_bbox"))
        comparisons = (
            ("image", pilot_row.get("image")),
            ("template", location.get("template")),
            ("question", location.get("question")),
            ("generated_answer", single_answer(location)),
            ("source_answer_text", location.get("source_answer_text")),
            ("source_bbox", expected_bbox),
        )
        for field, expected in comparisons:
            v.equal(record.get(field), expected,
                    f"review t4_location_review row {row_number}.{field}")

    sanity_sheet = sheet_records(
        "sanity_check", ("question_id", "sample_id", "claimed", "recomputed", "match")
    )
    sanity_ids = [str(record.get("question_id")) for record in sanity_sheet]
    v.check(not duplicates(sanity_ids),
            f"review sanity_check: duplicate question ids {duplicates(sanity_ids)}")
    v.check(set(sanity_ids) <= final_location_ids,
            f"review sanity_check: questions absent from final pilot "
            f"{sorted(set(sanity_ids) - final_location_ids)}")
    sanity_matches = sum(str(record.get("match", "")).lower() == "yes"
                         for record in sanity_sheet)
    sanity_meta = stats.get("sanity_check")
    if isinstance(sanity_meta, dict):
        v.equal(sanity_meta.get("sampled"), len(sanity_sheet),
                "meta.stats.sanity_check.sampled")
        v.equal(sanity_meta.get("matched"), sanity_matches,
                "meta.stats.sanity_check.matched")
    else:
        v.errors.append("meta.stats.sanity_check is missing or invalid")
    for row_number, record in enumerate(sanity_sheet, 2):
        v.equal(record.get("claimed"), record.get("recomputed"),
                f"review sanity_check row {row_number}: claimed vs recomputed")
        v.equal(str(record.get("match", "")).lower(), "yes",
                f"review sanity_check row {row_number}.match")

    audit_sheet = sheet_records(
        "visual_audit",
        AUDIT_SHEET_FIELDS,
        exact_headers=AUDIT_SHEET_FIELDS,
    )
    v.equal(len(audit_sheet), len(audit), "review visual_audit data rows")
    audit_sheet_ids = [str(record.get("question_id")) for record in audit_sheet]
    v.check(not duplicates(audit_sheet_ids),
            f"review visual_audit: duplicate question ids {duplicates(audit_sheet_ids)}")
    v.equal(set(audit_sheet_ids), set(audit_ids),
            "review visual_audit question-id set")
    audit_by_id = {str(record.get("question_id")): record for record in audit}
    for row_number, record in enumerate(audit_sheet, 2):
        qid = str(record.get("question_id"))
        source = audit_by_id.get(qid)
        if source is None:
            v.errors.append(f"review visual_audit row {row_number}: unknown question {qid}")
            continue
        for field in AUDIT_FIELDS:
            expected = source.get(field)
            if isinstance(expected, (dict, list)):
                expected = json.dumps(expected, ensure_ascii=False)
            v.equal(normalized_cell(record.get(field)), normalized_cell(expected),
                    f"review visual_audit row {row_number}.{field}")
        retained = qid in final_location_ids
        v.equal(record.get("retained_after_audit"), retained,
                f"review visual_audit row {row_number}.retained_after_audit")
        v.equal(retained, source.get("visual_verdict") == "pass",
                f"review visual_audit row {row_number}: retention vs verdict")

    # ------------------------------------------------------ optional blind A/B reviews
    pairs_sha256 = sha256_file(root / PAIRS_REL)
    for reviewer, relative_path in PAIR_REVIEW_RELS.items():
        review_path = root / relative_path
        if not review_path.exists():
            continue
        blind_workbook = load_workbook(review_path, read_only=True, data_only=False)
        workbook_label = f"blind review {reviewer}"
        v.equal(blind_workbook.sheetnames, ["pairs", "instructions"],
                f"{workbook_label}: sheet names/order")
        if "pairs" not in blind_workbook.sheetnames:
            continue
        pair_sheet = blind_workbook["pairs"]
        pair_rows = pair_sheet.iter_rows(values_only=True)
        try:
            pair_headers = [str(value) if value is not None else ""
                            for value in next(pair_rows)]
        except StopIteration:
            v.errors.append(f"{workbook_label}: pairs sheet is empty")
            continue
        v.equal(pair_headers, list(PAIR_REVIEW_COLUMNS),
                f"{workbook_label}: exact pairs header schema")
        leaked = set(pair_headers) & PAIR_REVIEW_PRIVATE_COLUMNS
        v.check(not leaked,
                f"{workbook_label}: private/draft columns leaked {sorted(leaked)}")
        blind_records = [dict(zip(pair_headers, values)) for values in pair_rows
                         if any(value is not None for value in values)]
        v.equal(len(blind_records), len(pairs), f"{workbook_label}: pair row count")
        v.equal([str(record.get("pair_id")) for record in blind_records], pair_ids,
                f"{workbook_label}: pair-id order")
        for row_number, (record, source) in enumerate(zip(blind_records, pairs), 2):
            for field in PAIR_REVIEW_COMMON_COLUMNS:
                v.equal(normalized_cell(record.get(field)),
                        normalized_cell(source.get(field)),
                        f"{workbook_label} row {row_number}.{field}")
            for field in PAIR_REVIEW_INPUT_COLUMNS:
                v.equal(normalized_cell(record.get(field)), "",
                        f"{workbook_label} row {row_number}.{field} must be blank")

        if "instructions" not in blind_workbook.sheetnames:
            continue
        instruction_sheet = blind_workbook["instructions"]
        instruction_values = list(instruction_sheet.iter_rows(values_only=True))
        if not instruction_values:
            v.errors.append(f"{workbook_label}: instructions sheet is empty")
            continue
        v.equal(list(instruction_values[0]), ["항목", "내용"],
                f"{workbook_label}: instructions header")
        instruction_keys = [str(row[0]) for row in instruction_values[1:]
                            if row and row[0] is not None]
        v.check(not duplicates(instruction_keys),
                f"{workbook_label}: duplicate instruction keys "
                f"{duplicates(instruction_keys)}")
        instructions = {
            str(row[0]): row[1] if len(row) > 1 else None
            for row in instruction_values[1:]
            if row and row[0] is not None
        }
        v.equal(instructions.get("reviewer"), reviewer,
                f"{workbook_label}: reviewer")
        v.equal(Path(str(instructions.get("source_manifest"))), PAIRS_REL,
                f"{workbook_label}: source_manifest")
        v.equal(instructions.get("source_sha256"), pairs_sha256,
                f"{workbook_label}: source_sha256")
        v.equal(instructions.get("pair_count"), len(pairs),
                f"{workbook_label}: pair_count")

    summary = {
        "screens": len(pilot),
        "content_questions": len(content_ids),
        "location_questions": len(location_ids),
        "pairs": len(pairs),
        "pending": pending_count,
        "draft_T4": pair_label_distribution.get("T4", 0),
        "draft_T2": pair_label_distribution.get("T2", 0),
        "draft_uncertain": pair_label_distribution.get("uncertain", 0),
        "visual_pass": verdict_counts.get("pass", 0),
        "visual_fail": verdict_counts.get("fail", 0),
    }
    return v, summary


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    default_root = Path(__file__).resolve().parents[2]
    parser.add_argument("--root", type=Path, default=default_root,
                        help="repository root (default: inferred from this script)")
    parser.add_argument("--max-errors", type=int, default=100,
                        help="maximum individual errors to print (default: 100)")
    args = parser.parse_args()
    if args.max_errors <= 0:
        parser.error("--max-errors must be positive")
    root = args.root.expanduser().resolve()

    try:
        validation, summary = validate(root)
    except (OSError, ValueError, RuntimeError, json.JSONDecodeError) as exc:
        print(f"T4 PILOT VALIDATION FATAL: {exc}", file=sys.stderr)
        return 2

    summary_text = ", ".join(f"{key}={value}" for key, value in summary.items())
    if validation.errors:
        print(f"T4 PILOT VALIDATION FAILED ({len(validation.errors)} errors)", file=sys.stderr)
        for error in validation.errors[:args.max_errors]:
            print(f"- {error}", file=sys.stderr)
        hidden = len(validation.errors) - args.max_errors
        if hidden > 0:
            print(f"- ... {hidden} additional errors omitted", file=sys.stderr)
        print(f"derived summary: {summary_text}")
        return 1

    print(f"T4 PILOT VALIDATION OK: {summary_text}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
